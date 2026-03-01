#!/usr/bin/env python3
"""
FGS App — Serveur avec base de données PostgreSQL
Utilise pg8000 (compatible Python 3.14+)
"""

import json, os, re, io, base64
import pg8000.native
from flask import Flask, request, jsonify, send_from_directory, send_file

app = Flask(__name__, static_folder=".")

# ── Connexion ─────────────────────────────────────────────────────────────────

def parse_db_url(url):
    """Parse une DATABASE_URL en paramètres de connexion."""
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)
    m = re.match(r'postgresql://([^:]+):([^@]+)@([^:/]+):?(\d*)/(.+)', url)
    if not m:
        raise ValueError(f"DATABASE_URL invalide : {url}")
    user, password, host, port, dbname = m.groups()
    return {
        "user": user,
        "password": password,
        "host": host,
        "port": int(port) if port else 5432,
        "database": dbname.split("?")[0],  # ignore ?sslmode=...
        "ssl_context": True,               # Render exige SSL
    }

def get_conn():
    url = os.environ.get("DATABASE_URL", "")
    if not url:
        raise RuntimeError("DATABASE_URL manquante.")
    p = parse_db_url(url)
    return pg8000.native.Connection(**p)

# ── Schéma et données par défaut ──────────────────────────────────────────────

SCHEMA_BD = """
CREATE TABLE IF NOT EXISTS fgs_data (
    id     SERIAL PRIMARY KEY,
    cle    TEXT UNIQUE NOT NULL,
    valeur TEXT NOT NULL,
    maj    TIMESTAMPTZ DEFAULT NOW()
)
"""

SCHEMA_POINTAGE = """
CREATE TABLE IF NOT EXISTS pointages (
    id          SERIAL PRIMARY KEY,
    date_jour   DATE NOT NULL,
    chantier    TEXT NOT NULL,
    auteur      TEXT NOT NULL,
    role_auteur TEXT NOT NULL DEFAULT 'chef',
    lignes      JSONB NOT NULL DEFAULT '[]',
    cree_le     TIMESTAMPTZ DEFAULT NOW(),
    maj_le      TIMESTAMPTZ DEFAULT NOW()
)
"""

SCHEMA_LIVCHANTIER = """
CREATE TABLE IF NOT EXISTS livraisons_chantier (
    id           SERIAL PRIMARY KEY,
    date_liv     DATE NOT NULL,
    chantier     TEXT NOT NULL,
    auteur       TEXT NOT NULL,
    element      TEXT NOT NULL,
    quantite     NUMERIC(12,3) NOT NULL DEFAULT 0,
    unite        TEXT NOT NULL DEFAULT 'm3',
    notes        TEXT DEFAULT '',
    cree_le      TIMESTAMPTZ DEFAULT NOW()
)
"""

SCHEMA_PRIX_REF = """
CREATE TABLE IF NOT EXISTS prix_reference (
    id        SERIAL PRIMARY KEY,
    chantier  TEXT NOT NULL,
    element   TEXT NOT NULL,
    prix_unitaire NUMERIC(12,2) NOT NULL,
    maj_le    TIMESTAMPTZ DEFAULT NOW(),
    UNIQUE(chantier, element)
)
"""

DONNEES_DEFAUT = {
    "categories": [
        {"id":"c1","nom":"Pelles","icone":"🦾"},
        {"id":"c2","nom":"Projection voie sèche","icone":"💨"},
        {"id":"c3","nom":"Injection","icone":"💉"},
        {"id":"c4","nom":"Manutention","icone":"🏗️"},
        {"id":"c5","nom":"Foreuses","icone":"🔩"},
        {"id":"c6","nom":"Glissières","icone":"🛤️"},
        {"id":"c7","nom":"Disqueuses thermiques","icone":"⚙️"},
    ],
    "machines": [
        {"id":"EQ-001","catId":"c1","nom":"Minipelle 2,5t","modele":"Kubota U25-3",
         "annee":2021,"heures":1620,"hEntretien":1500,"seuil":250,"vgp":"2026-12-03",
         "site":"Lyon","serie":"KB-U25-3-XYZ","poids":2500,
         "piecesAssociees":[
             {"pieceId":"P-001","heuresInstallation":1500,"dateInstallation":"2026-01-10"},
             {"pieceId":"P-003","heuresInstallation":1200,"dateInstallation":"2025-10-01"},
         ]},
        {"id":"EQ-002","catId":"c4","nom":"Chariot télescopique","modele":"Manitou MT625",
         "annee":2020,"heures":3100,"hEntretien":2900,"seuil":200,"vgp":"2026-04-15",
         "site":"Grenoble","piecesAssociees":[]},
    ],
    "pieces": [
        {"id":"P-001","nom":"Filtre huile moteur","ref":"FH-123","dureeVal":250,"dureeUnite":"heures","stock":3,"notes":"Utiliser pièce d'origine"},
        {"id":"P-002","nom":"Courroie de distribution","ref":"CR-321","dureeVal":500,"dureeUnite":"heures","stock":1},
        {"id":"P-003","nom":"Filtre à air","ref":"FA-456","dureeVal":500,"dureeUnite":"heures","stock":2},
        {"id":"P-004","nom":"Filtre hydraulique","ref":"FH-789","dureeVal":1000,"dureeUnite":"heures","stock":0},
    ],
    "interventions": [
        {"id":"I-001","machineId":"EQ-001","date":"2026-01-10","datePrevue":"","heures":1500,
         "type":"entretien","notes":"Vidange + filtre huile","piecesChangees":["P-001"]},
        {"id":"I-002","machineId":"EQ-002","date":"2026-02-01","datePrevue":"","heures":2900,
         "type":"entretien","notes":"Révision à 2900h","piecesChangees":[]},
    ],
    "livraisons": [],
}

def init_db():
    conn = get_conn()
    conn.run(SCHEMA_BD)
    conn.run(SCHEMA_POINTAGE)
    conn.run(SCHEMA_LIVCHANTIER)
    conn.run(SCHEMA_PRIX_REF)
    rows = conn.run("SELECT COUNT(*) FROM fgs_data WHERE cle = 'bd'")
    if rows[0][0] == 0:
        conn.run(
            "INSERT INTO fgs_data (cle, valeur) VALUES (:cle, :valeur)",
            cle="bd",
            valeur=json.dumps(DONNEES_DEFAUT, ensure_ascii=False)
        )
    conn.close()
    print("Base PostgreSQL initialisee (fgs_data + pointages).")

# ── Lecture / écriture ────────────────────────────────────────────────────────

def lire():
    conn = get_conn()
    rows = conn.run("SELECT valeur FROM fgs_data WHERE cle = 'bd'")
    conn.close()
    if not rows:
        init_db()
        return DONNEES_DEFAUT
    val = rows[0][0]
    return json.loads(val) if isinstance(val, str) else val

def ecrire(bd):
    conn = get_conn()
    conn.run(
        """
        INSERT INTO fgs_data (cle, valeur, maj) VALUES (:cle, :valeur, NOW())
        ON CONFLICT (cle) DO UPDATE SET valeur = EXCLUDED.valeur, maj = NOW()
        """,
        cle="bd",
        valeur=json.dumps(bd, ensure_ascii=False)
    )
    conn.close()

# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(".", "app.html")


def get_role(identifiant):
    """
    Retourne le rôle d'un utilisateur depuis ses variables d'environnement.
    Format de la variable : identifiant=motdepasse:role
    Rôles : admin, rh, chef
    Exemple : Admin=MonMDP:admin  /  rh.martin=MDP:rh  /  chef.lyon=MDP:chef
    Compatibilité ancien format (juste mot de passe sans rôle) = chef par défaut
    """
    for key, value in os.environ.items():
        if key.lower() == identifiant.lower():
            parts = value.split(":")
            if len(parts) >= 2:
                return parts[-1].strip().lower()  # dernier segment = rôle
            return "chef"  # ancien format sans rôle = chef par défaut
    return "chef"


@app.route("/api/whoami")
def whoami():
    """Debug: retourne les variables d'env disponibles (sans les valeurs) pour diagnostic"""
    comptes = []
    for key, value in os.environ.items():
        if key.lower() in ('admin', 'rh') or (len(key) < 20 and not key.startswith('_') and not key.startswith('RENDER') and not key.startswith('DATABASE') and not key.startswith('PORT') and not key.startswith('PATH') and not key.startswith('HOME') and not key.startswith('USER') and not key.startswith('PWD') and not key.startswith('LANG') and not key.startswith('LC_') and not key.startswith('PYTHON') and not key.startswith('PIP') and not key.startswith('VIRTUAL')):
            # Déduire le rôle comme auth() le ferait
            id_lower = key.lower()
            role_deduit = "admin" if id_lower == "admin" else ("rh" if id_lower == "rh" else "chef")
            has_colon = ":" in value
            if has_colon:
                last_colon = value.rfind(":")
                role_candidat = value[last_colon+1:].strip().lower()
                if role_candidat in ("chef", "admin", "rh"):
                    role_deduit = role_candidat
            comptes.append({"identifiant": key, "role_deduit": role_deduit, "format": "nouveau" if has_colon else "ancien"})
    return jsonify({"comptes": comptes})

@app.route("/api/auth", methods=["POST"])
def auth():
    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"ok": False}), 400
        identifiant = data.get("identifiant", "").strip()
        mdp = data.get("mdp", "")
        if not identifiant or not mdp:
            return jsonify({"ok": False}), 400
        for key, value in os.environ.items():
            if key.lower() != identifiant.lower():
                continue
            value = value.strip()
            # Format variable Render : "motdepasse:role" (ex: MonMDP:admin, MonMDP:chef)
            # ou "motdepasse" seul (ancien format sans rôle -> chef par défaut)
            # L'utilisateur saisit UNIQUEMENT le mot de passe, sans le :role
            role = "chef"
            mdp_stocke = value
            if ":" in value:
                last_colon = value.rfind(":")
                role_candidat = value[last_colon+1:].strip().lower()
                if role_candidat in ("chef", "admin", "rh"):
                    role = role_candidat
                    mdp_stocke = value[:last_colon]
                # sinon le ":" fait partie du mdp (ex: "Illite@8020") -> mdp_stocke = value entier
            if mdp_stocke == mdp:
                return jsonify({"ok": True, "nom": key, "role": role})
        return jsonify({"ok": False})
    except Exception as e:
        return jsonify({"ok": False, "erreur": str(e)}), 500



@app.route("/api/pointages", methods=["GET"])
def get_pointages():
    """Retourne les pointages selon le rôle."""
    try:
        role   = request.args.get("role", "chef")
        auteur = request.args.get("auteur", "")
        date_debut = request.args.get("debut", "")
        date_fin   = request.args.get("fin", "")

        conn = get_conn()
        if role in ("admin", "rh"):
            if date_debut and date_fin:
                rows = conn.run(
                    "SELECT * FROM pointages WHERE date_jour BETWEEN :d AND :f ORDER BY date_jour DESC, cree_le DESC",
                    d=date_debut, f=date_fin
                )
            else:
                rows = conn.run("SELECT * FROM pointages ORDER BY date_jour DESC, cree_le DESC LIMIT 200")
        else:
            # Chef : uniquement ses propres pointages
            if date_debut and date_fin:
                rows = conn.run(
                    "SELECT * FROM pointages WHERE auteur ILIKE :a AND date_jour BETWEEN :d AND :f ORDER BY date_jour DESC",
                    a=auteur, d=date_debut, f=date_fin
                )
            else:
                rows = conn.run(
                    "SELECT * FROM pointages WHERE auteur ILIKE :a ORDER BY date_jour DESC LIMIT 100",
                    a=auteur
                )
        conn.close()

        cols = ["id","date_jour","chantier","auteur","role_auteur","lignes","cree_le","maj_le"]
        result = []
        for row in rows:
            d = dict(zip(cols, row))
            d["date_jour"] = str(d["date_jour"])
            d["cree_le"]   = str(d["cree_le"])
            d["maj_le"]    = str(d["maj_le"])
            if isinstance(d["lignes"], str):
                d["lignes"] = json.loads(d["lignes"])
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/pointages", methods=["POST"])
def save_pointage():
    """Crée ou met à jour un pointage."""
    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"erreur": "Données invalides"}), 400

        pid        = data.get("id")
        date_jour  = data.get("date_jour")
        chantier   = data.get("chantier", "")
        auteur     = data.get("auteur", "")
        role_auteur= data.get("role_auteur", "chef")
        lignes     = data.get("lignes", [])

        conn = get_conn()
        if pid:
            conn.run(
                "UPDATE pointages SET lignes=:l, chantier=:c, maj_le=NOW() WHERE id=:id",
                l=json.dumps(lignes, ensure_ascii=False), c=chantier, id=pid
            )
            new_id = pid
        else:
            rows = conn.run(
                """INSERT INTO pointages (date_jour, chantier, auteur, role_auteur, lignes)
                   VALUES (:d, :c, :a, :r, :l) RETURNING id""",
                d=date_jour, c=chantier, a=auteur, r=role_auteur,
                l=json.dumps(lignes, ensure_ascii=False)
            )
            new_id = rows[0][0]
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/pointages/<int:pid>", methods=["DELETE"])
def delete_pointage(pid):
    try:
        role = request.args.get("role", "chef")
        auteur = request.args.get("auteur", "")
        conn = get_conn()
        if role in ("admin", "rh"):
            conn.run("DELETE FROM pointages WHERE id=:id", id=pid)
        else:
            conn.run("DELETE FROM pointages WHERE id=:id AND auteur ILIKE :a", id=pid, a=auteur)
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500




@app.route("/api/bd", methods=["GET"])
def get_bd():
    try:
        return jsonify(lire())
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500

@app.route("/api/bd", methods=["POST"])
def post_bd():
    try:
        bd = request.get_json(force=True, silent=True)
        if bd is None:
            return jsonify({"erreur": "JSON invalide"}), 400
        ecrire(bd)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/backup")
def backup():
    """
    Télécharge une sauvegarde complète des données au format JSON.
    Protégé par une clé secrète : /api/backup?key=VOTRE_CLE
    
    Sur Render : ajoutez une variable d'environnement BACKUP_KEY = votre_mot_de_passe_secret
    """
    cle = request.args.get("key", "")
    cle_attendue = os.environ.get("BACKUP_KEY", "")
    
    if not cle_attendue:
        return jsonify({"erreur": "Variable BACKUP_KEY non configurée sur Render."}), 500
    if cle != cle_attendue:
        return jsonify({"erreur": "Clé incorrecte."}), 403
    
    try:
        bd = lire()
        from datetime import datetime
        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        nom_fichier = f"fgs-backup-{date_str}.json"
        contenu = json.dumps(bd, ensure_ascii=False, indent=2)
        
        from flask import Response
        return Response(
            contenu,
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename={nom_fichier}"}
        )
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/restore", methods=["POST"])
def restore():
    """
    Restaure les données depuis un fichier JSON.
    Protégé par la même clé secrète BACKUP_KEY.
    
    Utilisation : POST /api/restore?key=VOTRE_CLE
    Corps : le fichier JSON de sauvegarde
    """
    cle = request.args.get("key", "")
    cle_attendue = os.environ.get("BACKUP_KEY", "")
    
    if not cle_attendue:
        return jsonify({"erreur": "Variable BACKUP_KEY non configurée sur Render."}), 500
    if cle != cle_attendue:
        return jsonify({"erreur": "Clé incorrecte."}), 403
    
    try:
        bd = request.get_json(force=True, silent=True)
        if not bd:
            return jsonify({"erreur": "Fichier JSON invalide ou vide."}), 400
        
        # Vérification basique que c'est bien une sauvegarde FGS
        champs_requis = ["categories", "machines", "pieces", "interventions", "livraisons"]
        for champ in champs_requis:
            if champ not in bd:
                return jsonify({"erreur": f"Sauvegarde invalide : champ '{champ}' manquant."}), 400
        
        ecrire(bd)
        return jsonify({
            "ok": True,
            "message": f"Restauration réussie : {len(bd.get('machines',[]))} machines, {len(bd.get('pieces',[]))} pièces, {len(bd.get('interventions',[]))} interventions."
        })
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/livchantier", methods=["GET"])
def get_livchantier():
    try:
        role    = request.args.get("role", "chef")
        auteur  = request.args.get("auteur", "")
        chantier= request.args.get("chantier", "")
        element = request.args.get("element", "")
        chef_f  = request.args.get("chef", "")
        debut   = request.args.get("debut", "")
        fin     = request.args.get("fin", "")

        where = []
        params = {}
        if role not in ("admin", "rh"):
            where.append("auteur ILIKE :auteur")
            params["auteur"] = auteur
        if chantier:
            where.append("chantier ILIKE :chantier")
            params["chantier"] = f"%{chantier}%"
        if element:
            where.append("element = :element")
            params["element"] = element
        if chef_f:
            where.append("auteur ILIKE :chef_f")
            params["chef_f"] = f"%{chef_f}%"
        if debut and fin:
            where.append("date_liv BETWEEN :debut AND :fin")
            params["debut"] = debut
            params["fin"] = fin

        sql = """SELECT l.id, l.date_liv, l.chantier, l.auteur, l.element,
                        l.quantite, l.unite, p.prix_unitaire, l.notes, l.cree_le
                 FROM livraisons_chantier l
                 LEFT JOIN prix_reference p ON p.chantier=l.chantier AND p.element=l.element"""
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY date_liv DESC, cree_le DESC"

        conn = get_conn()
        rows = conn.run(sql, **params) if params else conn.run(sql)
        conn.close()

        cols = ["id","date_liv","chantier","auteur","element","quantite","unite","prix_unitaire","notes","cree_le"]
        result = []
        for row in rows:
            d = dict(zip(cols, row))
            d["date_liv"] = str(d["date_liv"])
            d["cree_le"]  = str(d["cree_le"])
            d["quantite"] = float(d["quantite"]) if d["quantite"] is not None else 0
            d["prix_unitaire"] = float(d["prix_unitaire"]) if d["prix_unitaire"] is not None else None
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/livchantier", methods=["POST"])
def save_livchantier():
    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"erreur": "Données invalides"}), 400
        lid      = data.get("id")
        date_liv = data.get("date_liv")
        chantier = data.get("chantier", "")
        auteur   = data.get("auteur", "")
        element  = data.get("element", "béton")
        quantite = float(data.get("quantite", 0))
        unite    = data.get("unite", "m3")
        notes    = data.get("notes", "")
        conn = get_conn()
        if lid:
            conn.run(
                """UPDATE livraisons_chantier
                   SET date_liv=:d, chantier=:c, element=:e, quantite=:q, unite=:u, notes=:n
                   WHERE id=:id""",
                d=date_liv, c=chantier, e=element, q=quantite, u=unite, n=notes, id=lid
            )
            new_id = lid
        else:
            rows = conn.run(
                """INSERT INTO livraisons_chantier
                   (date_liv,chantier,auteur,element,quantite,unite,notes)
                   VALUES (:d,:c,:a,:e,:q,:u,:n) RETURNING id""",
                d=date_liv, c=chantier, a=auteur, e=element, q=quantite, u=unite, n=notes
            )
            new_id = rows[0][0]
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/livchantier/<int:lid>", methods=["DELETE"])
def delete_livchantier(lid):
    try:
        role   = request.args.get("role", "chef")
        auteur = request.args.get("auteur", "")
        conn = get_conn()
        if role in ("admin", "rh"):
            conn.run("DELETE FROM livraisons_chantier WHERE id=:id", id=lid)
        else:
            conn.run("DELETE FROM livraisons_chantier WHERE id=:id AND auteur ILIKE :a",
                     id=lid, a=auteur)
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/prix_reference", methods=["GET"])
def get_prix_ref():
    """Retourne tous les prix de référence (chantier x element)."""
    try:
        conn = get_conn()
        rows = conn.run("SELECT id, chantier, element, prix_unitaire, maj_le FROM prix_reference ORDER BY chantier, element")
        conn.close()
        cols = ["id","chantier","element","prix_unitaire","maj_le"]
        result = []
        for row in rows:
            d = dict(zip(cols, row))
            d["prix_unitaire"] = float(d["prix_unitaire"]) if d["prix_unitaire"] else None
            d["maj_le"] = str(d["maj_le"])
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/prix_reference", methods=["POST"])
def save_prix_ref():
    """Upsert un prix de référence pour un couple chantier+element."""
    try:
        data = request.get_json(force=True, silent=True)
        chantier = data.get("chantier", "").strip()
        element  = data.get("element", "").strip()
        prix     = float(data.get("prix_unitaire", 0))
        if not chantier or not element:
            return jsonify({"erreur": "chantier et element requis"}), 400
        conn = get_conn()
        conn.run(
            """INSERT INTO prix_reference (chantier, element, prix_unitaire, maj_le)
               VALUES (:c, :e, :p, NOW())
               ON CONFLICT (chantier, element)
               DO UPDATE SET prix_unitaire=EXCLUDED.prix_unitaire, maj_le=NOW()""",
            c=chantier, e=element, p=prix
        )
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500

@app.route("/api/sante")
def sante():
    try:
        conn = get_conn()
        rows = conn.run("SELECT maj FROM fgs_data WHERE cle='bd'")
        conn.close()
        derniere_maj = str(rows[0][0]) if rows else "jamais"
        return jsonify({"ok": True, "derniere_maj": derniere_maj})
    except Exception as e:
        return jsonify({"ok": False, "erreur": str(e)}), 500

# ── Démarrage ─────────────────────────────────────────────────────────────────

# ── Initialisation au démarrage (fonctionne avec Gunicorn ET python direct) ──
# Appelé à l'import du module — crée la table si elle n'existe pas
try:
    init_db()
except Exception as e:
    print(f"AVERTISSEMENT init_db: {e}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"Serveur demarre sur http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)


# ── Anomalies ─────────────────────────────────────────────────────────────────

SCHEMA_ANOMALIES = """
CREATE TABLE IF NOT EXISTS anomalies (
    id          SERIAL PRIMARY KEY,
    date_ano    DATE NOT NULL DEFAULT CURRENT_DATE,
    nom_machine TEXT NOT NULL,
    num_parc    TEXT NOT NULL DEFAULT '',
    auteur      TEXT NOT NULL,
    description TEXT NOT NULL,
    statut      TEXT NOT NULL DEFAULT 'ouvert',
    cree_le     TIMESTAMPTZ DEFAULT NOW()
)
"""

def init_anomalies():
    conn = get_conn()
    conn.run(SCHEMA_ANOMALIES)
    conn.close()

try:
    init_anomalies()
except Exception as e:
    print(f"AVERTISSEMENT init_anomalies: {e}")


@app.route("/api/anomalies", methods=["GET"])
def get_anomalies():
    try:
        role   = request.args.get("role", "chef")
        auteur = request.args.get("auteur", "")
        conn   = get_conn()
        if role in ("admin", "rh"):
            rows = conn.run(
                "SELECT id,date_ano,nom_machine,num_parc,auteur,description,statut,cree_le "
                "FROM anomalies ORDER BY cree_le DESC LIMIT 200"
            )
        else:
            rows = conn.run(
                "SELECT id,date_ano,nom_machine,num_parc,auteur,description,statut,cree_le "
                "FROM anomalies WHERE auteur ILIKE :a ORDER BY cree_le DESC LIMIT 100",
                a=auteur
            )
        conn.close()
        cols = ["id","date_ano","nom_machine","num_parc","auteur","description","statut","cree_le"]
        result = []
        for row in rows:
            d = dict(zip(cols, row))
            d["date_ano"] = str(d["date_ano"])
            d["cree_le"]  = str(d["cree_le"])
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/anomalies", methods=["POST"])
def save_anomalie():
    try:
        data        = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"erreur": "Données invalides"}), 400
        aid         = data.get("id")
        date_ano    = data.get("date_ano", "")
        nom_machine = data.get("nom_machine", "").strip()
        num_parc    = data.get("num_parc", "").strip()
        auteur      = data.get("auteur", "")
        description = data.get("description", "").strip()
        statut      = data.get("statut", "ouvert")
        if not nom_machine or not description:
            return jsonify({"erreur": "Champs obligatoires manquants"}), 400
        conn = get_conn()
        if aid:
            conn.run(
                "UPDATE anomalies SET statut=:s, description=:d WHERE id=:id",
                s=statut, d=description, id=aid
            )
            new_id = aid
        else:
            rows = conn.run(
                """INSERT INTO anomalies (date_ano,nom_machine,num_parc,auteur,description,statut)
                   VALUES (:da,:nm,:np,:au,:de,:st) RETURNING id""",
                da=date_ano, nm=nom_machine, np=num_parc, au=auteur, de=description, st=statut
            )
            new_id = rows[0][0]
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/anomalies/<int:aid>", methods=["DELETE"])
def delete_anomalie(aid):
    try:
        role   = request.args.get("role", "chef")
        auteur = request.args.get("auteur", "")
        conn   = get_conn()
        if role in ("admin", "rh"):
            conn.run("DELETE FROM anomalies WHERE id=:id", id=aid)
        else:
            conn.run("DELETE FROM anomalies WHERE id=:id AND auteur ILIKE :a", id=aid, a=auteur)
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/anomalies/<int:aid>/statut", methods=["POST"])
def update_statut_anomalie(aid):
    """Admin peut changer le statut : ouvert -> en_cours -> resolu"""
    try:
        data   = request.get_json(force=True, silent=True)
        statut = data.get("statut", "ouvert")
        conn   = get_conn()
        conn.run("UPDATE anomalies SET statut=:s WHERE id=:id", s=statut, id=aid)
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# EXPORTS XLSX PROFESSIONNELS — openpyxl avec logo FGS
# ══════════════════════════════════════════════════════════════════════════════

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles.numbers import FORMAT_NUMBER_00
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Constantes couleurs ────────────────────────────────────────────────────────
_ROUGE      = "C0392B"
_ROUGE_L    = "FDECEA"
_GRIS_FOND  = "1A1A2E"
_BLEU_ACIER = "2C3E50"
_GRIS_CLAIR = "F0F0F5"
_BLANC      = "FFFFFF"
_BORDURE    = "D0D0DC"
_ORANGE     = "E67E22"
_VERT       = "1E8449"

# ── Helpers styles ─────────────────────────────────────────────────────────────
def _side(color=_BORDURE, style='thin'):
    return Side(style=style, color=color)

def _fill(color):
    return PatternFill(fill_type='solid', fgColor=color)

def _font(size=10, bold=False, color='000000', italic=False, name='Calibri'):
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)

def _align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _border_full(color=_BORDURE):
    s = _side(color)
    return Border(left=s, right=s, top=s, bottom=s)

def _border_bottom(color=_BORDURE):
    return Border(bottom=_side(color), right=_side(color))

# ── Cartouche professionnel ────────────────────────────────────────────────────
def _cartouche(ws, titre, sous_titre, auteur, nb_cols):
    """Construit le cartouche FGS sur 6 lignes. Retourne le numéro de la 1ère ligne de données."""
    from datetime import datetime
    date_str = datetime.now().strftime("%d/%m/%Y à %Hh%M")

    ws.row_dimensions[1].height = 6    # marge
    ws.row_dimensions[2].height = 36   # société + logo
    ws.row_dimensions[3].height = 30   # titre document (fond rouge)
    ws.row_dimensions[4].height = 20   # sous-titre
    ws.row_dimensions[5].height = 16   # date + exporteur
    ws.row_dimensions[6].height = 6    # séparateur

    # Remplir fond gris foncé lignes 2, 4, 5
    for row in (2, 4, 5):
        for col in range(1, nb_cols + 1):
            ws.cell(row=row, column=col).fill = _fill(_GRIS_FOND)

    # Ligne rouge (titre)
    for col in range(1, nb_cols + 1):
        ws.cell(row=3, column=col).fill = _fill(_ROUGE)

    # Séparateur clair
    for col in range(1, nb_cols + 1):
        ws.cell(row=6, column=col).fill = _fill(_GRIS_CLAIR)

    # Ligne 2 : Nom société
    c = ws.cell(row=2, column=1, value="FGS TRAVAUX SPÉCIAUX")
    c.font = _font(18, bold=True, color=_BLANC)
    c.alignment = _align('left', 'center')

    # Ligne 3 : Titre document
    c = ws.cell(row=3, column=1, value=titre.upper())
    c.font = _font(14, bold=True, color=_BLANC)
    c.alignment = _align('left', 'center')

    # Ligne 4 : Sous-titre
    c = ws.cell(row=4, column=1, value=sous_titre)
    c.font = _font(10, italic=True, color='CCCCCC')
    c.alignment = _align('left', 'center')

    # Ligne 5 : Date + auteur
    c = ws.cell(row=5, column=1, value=f"Exporté le {date_str}   •   Par : {auteur}")
    c.font = _font(9, color='AAAAAA')
    c.alignment = _align('left', 'center')

    # Logo
    logo_path = os.path.join(os.path.dirname(__file__), 'logo_fgs.png')
    if os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            ratio = img.width / max(img.height, 1)
            img.height = 80
            img.width  = int(80 * ratio)
            col_logo = max(nb_cols - 1, 2)
            img.anchor = f"{get_column_letter(col_logo)}2"
            ws.add_image(img)
        except Exception:
            pass

    return 7  # première ligne de données

def _entetes(ws, row, cols, hauteur=22):
    """En-têtes tableau avec fond bleu acier."""
    ws.row_dimensions[row].height = hauteur
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.font  = _font(10, bold=True, color=_BLANC)
        c.fill  = _fill(_BLEU_ACIER)
        c.alignment = _align('center', 'center')
        c.border = Border(
            left=_side(_BLANC, 'thin'), right=_side(_BLANC, 'thin'),
            top=_side(_BLANC, 'thin'), bottom=_side(_BLANC, 'thin')
        )

def _ligne(ws, row, data, idx=0, hauteur=18, surbrillance=None):
    """Ligne de données alternée."""
    ws.row_dimensions[row].height = hauteur
    if surbrillance:
        bg = surbrillance
    else:
        bg = _GRIS_CLAIR if idx % 2 == 0 else _BLANC
    for i, val in enumerate(data, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font  = _font(10)
        c.fill  = _fill(bg)
        c.alignment = _align('left', 'center', wrap=True)
        c.border = _border_bottom()

def _sous_titre_section(ws, row, texte, nb_cols, hauteur=16):
    """Ligne de section (fond gris clair, texte gras)."""
    ws.row_dimensions[row].height = hauteur
    for col in range(1, nb_cols + 1):
        ws.cell(row=row, column=col).fill = _fill(_GRIS_CLAIR)
    c = ws.cell(row=row, column=1, value=texte)
    c.font = _font(10, bold=True, color=_BLEU_ACIER)
    c.alignment = _align('left', 'center')

def _col_widths(ws, largeurs):
    for i, w in enumerate(largeurs, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _make_wb():
    wb = Workbook()
    wb.remove(wb.active)
    return wb

def _send_wb(wb, nom_fichier):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f"{nom_fichier}.xlsx"
    )

# ── Route principale d'export ──────────────────────────────────────────────────
@app.route("/api/export/<type_export>", methods=["POST"])
def export_xlsx(type_export):
    if not OPENPYXL_OK:
        return jsonify({"erreur": "openpyxl non installé sur le serveur"}), 500
    try:
        payload = request.get_json(force=True, silent=True) or {}
        auteur = payload.get("auteur", "—")
        bd     = payload.get("bd", {})
        data   = payload.get("data", {})
        from datetime import datetime
        date_iso = datetime.now().strftime("%Y-%m-%d")

        if type_export == "parc":
            return _export_parc(bd, auteur, date_iso)
        elif type_export == "entretiens":
            return _export_entretiens(bd, auteur, date_iso)
        elif type_export == "pieces":
            return _export_pieces(bd, auteur, date_iso)
        elif type_export == "pointage_jour":
            return _export_pointage_jour(data.get("pointages", []), auteur, date_iso)
        elif type_export == "pointage_semaine":
            return _export_pointage_semaine(data.get("pointages", []), auteur, date_iso)
        elif type_export == "livraisons_admin":
            return _export_livraisons_admin(data, auteur, date_iso)
        elif type_export == "livraisons_chef":
            return _export_livraisons_chef(data, auteur, date_iso)
        elif type_export == "total_chantier":
            return _export_total_chantier(data, bd, auteur, date_iso)
        else:
            return jsonify({"erreur": f"Type inconnu : {type_export}"}), 400
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"erreur": str(e)}), 500


# ── Export Parc matériel ───────────────────────────────────────────────────────
def _export_parc(bd, auteur, date_iso):
    machines    = bd.get("machines", [])
    categories  = bd.get("categories", [])

    def ncat(cid):
        for c in categories:
            if c.get("id") == cid: return c.get("nom", "—")
        return "—"

    def st_mach(m):
        seuil = m.get("seuil", 250)
        d = (m.get("heures", 0) or 0) - (m.get("hEntretien", 0) or 0)
        pct = d / max(seuil, 1)
        vgp = m.get("vgp", "")
        if vgp:
            from datetime import date
            try:
                days = (date.fromisoformat(vgp) - date.today()).days
                if days < 0: return "En retard"
                if days < 30: return "À prévoir"
            except: pass
        if pct >= 1: return "En retard"
        if pct >= 0.8: return "À prévoir"
        return "À jour"

    cols = ["Machine","Catégorie","Modèle","Année","Site / Chantier",
            "Heures","Dernier entretien","Écart / Seuil","Statut","VGP prochaine"]
    largeurs = [28, 18, 20, 10, 18, 12, 16, 16, 12, 16]

    wb = _make_wb()
    ws = wb.create_sheet("Parc matériel")
    _col_widths(ws, largeurs)
    r = _cartouche(ws, "Parc Matériel", f"{len(machines)} engin(s) enregistré(s)", auteur, len(cols))
    _entetes(ws, r, cols); r += 1

    # Grouper par catégorie
    cats_used = []
    cat_dict = {}
    for m in machines:
        cid = m.get("catId", "—")
        cn  = ncat(cid)
        if cn not in cat_dict: cat_dict[cn] = []
        cat_dict[cn].append(m)
        if cn not in cats_used: cats_used.append(cn)

    idx = 0
    for cn in cats_used:
        _sous_titre_section(ws, r, f"  {cn}", len(cols)); r += 1
        for m in cat_dict[cn]:
            d = (m.get("heures", 0) or 0) - (m.get("hEntretien", 0) or 0)
            seuil = m.get("seuil", 250)
            st = st_mach(m)
            surb = None
            if st == "En retard": surb = _ROUGE_L
            elif st == "À prévoir": surb = "FEF9C3"
            vgp = m.get("vgp", "") or "—"
            if vgp != "—":
                try:
                    from datetime import date
                    vgp = date.fromisoformat(vgp).strftime("%d/%m/%Y")
                except: pass
            row_data = [
                m.get("nom","—"), cn, m.get("modele","—") or "—",
                m.get("annee","—") or "—", m.get("site","—") or "—",
                f"{m.get('heures',0)} h", f"{m.get('hEntretien',0)} h",
                f"{d} h / {seuil} h", st, vgp
            ]
            _ligne(ws, r, row_data, idx, surbrillance=surb); idx += 1; r += 1

    # Figer le cartouche
    ws.freeze_panes = f"A8"
    return _send_wb(wb, f"FGS-Parc-{date_iso}")


# ── Export Entretiens ──────────────────────────────────────────────────────────
def _export_entretiens(bd, auteur, date_iso):
    interventions = sorted(bd.get("interventions", []), key=lambda x: x.get("date",""), reverse=True)
    machines  = bd.get("machines", [])
    pieces    = bd.get("pieces", [])
    categories = bd.get("categories", [])

    def get_machine(mid):
        for m in machines:
            if m.get("id") == mid: return m
        return {}
    def get_piece(pid):
        for p in pieces:
            if p.get("id") == pid: return p
        return {}
    def ncat(cid):
        for c in categories:
            if c.get("id") == cid: return c.get("nom","—")
        return "—"
    def lbl_ty(t):
        return {"entretien":"Entretien préventif","reparation":"Réparation",
                "VGP":"VGP","remplacement":"Remplacement pièce"}.get(t, t)

    cols = ["Date","Machine","Catégorie","Site","Type","Heures",
            "Pièces remplacées","Pièces libres","Notes"]
    largeurs = [14, 24, 18, 16, 20, 10, 28, 24, 32]
    COULEURS = {
        "entretien": "EAF7EA",
        "reparation": _ROUGE_L,
        "VGP": "EAF0FB",
        "remplacement": "FEF3DC",
    }

    wb = _make_wb()
    ws = wb.create_sheet("Entretiens")
    _col_widths(ws, largeurs)
    r = _cartouche(ws, "Historique des Entretiens", f"{len(interventions)} intervention(s)", auteur, len(cols))
    _entetes(ws, r, cols); r += 1

    for idx, iv in enumerate(interventions):
        m = get_machine(iv.get("machineId",""))
        pcs = [get_piece(pid).get("nom","?") for pid in (iv.get("piecesChangees") or []) if get_piece(pid)]
        autres = [f"{ap.get('nom','?')} ({ap.get('ref','—')})" for ap in (iv.get("autresPieces") or [])]
        ty = iv.get("type","entretien")
        surb = COULEURS.get(ty)
        date_fmt = iv.get("date","—")
        if date_fmt and date_fmt != "—":
            try:
                from datetime import date
                date_fmt = date.fromisoformat(date_fmt).strftime("%d/%m/%Y")
            except: pass
        row_data = [
            date_fmt,
            m.get("nom","—"),
            ncat(m.get("catId","")),
            m.get("site","—") or "—",
            lbl_ty(ty),
            f"{iv.get('heures','—')} h" if iv.get("heures") else "—",
            ", ".join(pcs) if pcs else "—",
            ", ".join(autres) if autres else "—",
            iv.get("notes","—") or "—"
        ]
        _ligne(ws, r, row_data, idx, surbrillance=surb); r += 1

    ws.freeze_panes = "A8"
    return _send_wb(wb, f"FGS-Entretiens-{date_iso}")


# ── Export Pièces ──────────────────────────────────────────────────────────────
def _export_pieces(bd, auteur, date_iso):
    pieces    = bd.get("pieces", [])
    machines  = bd.get("machines", [])

    def get_machines_compat(p):
        ids = p.get("machinesCompatibles") or []
        if not ids:
            # fallback legacy
            ids = [m.get("id") for m in machines if any(pa.get("pieceId") == p.get("id") for pa in (m.get("piecesAssociees") or []))]
        return [m.get("nom","?") for m in machines if m.get("id") in ids]

    cols = ["Nom pièce","Référence","Durée de vie","Unité",
            "Stock actuel","Machines compatibles","Besoin (1/machine)","Statut stock","Notes"]
    largeurs = [24, 16, 14, 10, 14, 36, 16, 16, 28]

    wb = _make_wb()
    ws = wb.create_sheet("Pièces")
    _col_widths(ws, largeurs)
    r = _cartouche(ws, "Catalogue Pièces & Stocks", f"{len(pieces)} référence(s)", auteur, len(cols))
    _entetes(ws, r, cols); r += 1

    for idx, p in enumerate(pieces):
        compat = get_machines_compat(p)
        besoin = len(compat)
        stock  = p.get("stock", 0) or 0
        manque = max(besoin - stock, 0)
        if besoin == 0: statut = "Non rattachée"
        elif manque > 0: statut = f"⚠ {manque} à commander"
        else: statut = "✓ Stock OK"
        surb = _ROUGE_L if manque > 0 else ("F0FFF4" if besoin > 0 else None)
        row_data = [
            p.get("nom","—"), p.get("ref","—"),
            p.get("dureeVal","—"), p.get("dureeUnite","heures"),
            stock, ", ".join(compat) if compat else "—",
            besoin if besoin else "—", statut,
            p.get("notes","—") or "—"
        ]
        _ligne(ws, r, row_data, idx, surbrillance=surb); r += 1

    ws.freeze_panes = "A8"
    return _send_wb(wb, f"FGS-Pieces-{date_iso}")


# ── Export Pointage journalier ─────────────────────────────────────────────────
def _export_pointage_jour(pointages, auteur, date_iso):
    sorted_ptg = sorted(pointages, key=lambda x: x.get("date_jour",""), reverse=True)

    cols = ["Date","Chantier","Saisi par","Employé","Heures","Gd Déplacement","Panier Repas","Notes"]
    largeurs = [16, 22, 18, 22, 10, 16, 14, 30]

    wb = _make_wb()
    ws = wb.create_sheet("Pointage journalier")
    _col_widths(ws, largeurs)
    r = _cartouche(ws, "Pointages — Détail Journalier", f"{len(pointages)} journée(s) saisie(s)", auteur, len(cols))
    _entetes(ws, r, cols); r += 1

    idx = 0
    for ptg in sorted_ptg:
        date_fmt = ptg.get("date_jour","—")
        try:
            from datetime import date
            d = date.fromisoformat(date_fmt)
            jours = ["Lundi","Mardi","Mercredi","Jeudi","Vendredi","Samedi","Dimanche"]
            date_fmt = f"{jours[d.weekday()]} {d.strftime('%d/%m/%Y')}"
        except: pass
        chantier = ptg.get("chantier","—")
        saisie_par = ptg.get("auteur","—")
        lignes = ptg.get("lignes") or []
        notes = ptg.get("notes","") or "—"

        if not lignes:
            _ligne(ws, r, [date_fmt, chantier, saisie_par, "—","—","—","—",notes], idx); idx+=1; r+=1
        else:
            for j, l in enumerate(lignes):
                row_data = [
                    date_fmt if j == 0 else "",
                    chantier if j == 0 else "",
                    saisie_par if j == 0 else "",
                    l.get("nom","—"),
                    l.get("heures","—"),
                    "✓" if l.get("gd") else "—",
                    "✓" if l.get("panier") else "—",
                    notes if j == 0 else ""
                ]
                _ligne(ws, r, row_data, idx); idx+=1; r+=1

        # Sous-total de la journée
        total_h = sum(float(l.get("heures",0) or 0) for l in lignes)
        nb_gd   = sum(1 for l in lignes if l.get("gd"))
        nb_pan  = sum(1 for l in lignes if l.get("panier"))
        ws.row_dimensions[r].height = 16
        c = ws.cell(row=r, column=1, value=f"Sous-total : {len(lignes)} présence(s) · {total_h:.1f}h · {nb_gd} GD · {nb_pan} paniers")
        c.font = _font(9, bold=True, color=_BLEU_ACIER, italic=True)
        c.fill = _fill(_GRIS_CLAIR)
        c.alignment = _align('left', 'center')
        for col in range(2, len(cols)+1):
            ws.cell(row=r, column=col).fill = _fill(_GRIS_CLAIR)
        r += 1
        # Ligne vide séparation
        ws.row_dimensions[r].height = 6; r += 1

    ws.freeze_panes = "A8"
    return _send_wb(wb, f"FGS-Pointage-Jour-{date_iso}")


# ── Export Pointage hebdomadaire ───────────────────────────────────────────────
def _export_pointage_semaine(pointages, auteur, date_iso):
    from datetime import date, timedelta

    # Grouper par semaine
    sem_map = {}
    for ptg in pointages:
        dj = ptg.get("date_jour","")
        try:
            d = date.fromisoformat(dj)
            lundi = d - timedelta(days=d.weekday())
            wk = lundi.isoformat()
        except: wk = dj[:7]
        if wk not in sem_map: sem_map[wk] = []
        sem_map[wk].append(ptg)

    wb = _make_wb()
    jours = ["Lun","Mar","Mer","Jeu","Ven","Sam","Dim"]
    cols = ["Employé","Chantier"] + jours + ["Total h","GD","Paniers"]
    largeurs = [24, 22, 8, 8, 8, 8, 8, 8, 8, 12, 10, 10]

    for wk in sorted(sem_map.keys(), reverse=True):
        ptgs = sem_map[wk]
        try:
            lun = date.fromisoformat(wk)
            dim = lun + timedelta(days=6)
            lbl = f"{lun.strftime('%d/%m')} - {dim.strftime('%d/%m/%Y')}"
            nom_feuille = f"Sem {lun.strftime('%d-%m')}"[:31]
        except:
            lbl = wk; nom_feuille = wk[:31]

        ws = wb.create_sheet(nom_feuille)
        _col_widths(ws, largeurs)
        r = _cartouche(ws, "Pointages Hebdomadaires",
                       f"Semaine du {lbl}  —  {len(ptgs)} journée(s)", auteur, len(cols))
        _entetes(ws, r, cols); r += 1

        # Construire grille employé x jours par chantier
        emp_map = {}
        for ptg in ptgs:
            dj = ptg.get("date_jour","")
            try: dow = date.fromisoformat(dj).weekday()
            except: dow = 0
            for l in (ptg.get("lignes") or []):
                k = f"{l.get('nom','?')}||{ptg.get('chantier','?')}"
                if k not in emp_map:
                    emp_map[k] = {"nom": l.get("nom","?"), "chantier": ptg.get("chantier","?"),
                                  "h": [0]*7, "gd": 0, "pan": 0}
                emp_map[k]["h"][dow] += float(l.get("heures",0) or 0)
                if l.get("gd"):  emp_map[k]["gd"] += 1
                if l.get("panier"): emp_map[k]["pan"] += 1

        idx = 0
        for e in sorted(emp_map.values(), key=lambda x: x["nom"]):
            tot = sum(e["h"])
            row_data = ([e["nom"], e["chantier"]] +
                       [round(h,1) if h>0 else "" for h in e["h"]] +
                       [round(tot,1), e["gd"] or "", e["pan"] or ""])
            _ligne(ws, r, row_data, idx); idx+=1; r+=1

        # Ligne totaux
        ws.row_dimensions[r].height = 20
        for col in range(1, len(cols)+1):
            ws.cell(row=r, column=col).fill = _fill(_BLEU_ACIER)
        ws.cell(row=r, column=1, value="TOTAL SEMAINE").font = _font(10, bold=True, color=_BLANC)
        ws.cell(row=r, column=1).alignment = _align('left','center')
        all_h = [sum(e["h"][i] for e in emp_map.values()) for i in range(7)]
        for i, h in enumerate(all_h):
            c = ws.cell(row=r, column=3+i, value=round(h,1) if h>0 else "")
            c.font = _font(10, bold=True, color=_BLANC)
            c.alignment = _align('center','center')
        total_sem = sum(all_h)
        c = ws.cell(row=r, column=10, value=round(total_sem,1))
        c.font = _font(10, bold=True, color=_BLANC)
        c.alignment = _align('center','center')
        r += 1

        ws.freeze_panes = "A8"

    return _send_wb(wb, f"FGS-Pointage-Semaine-{date_iso}")


# ── Export Livraisons admin ────────────────────────────────────────────────────
def _export_livraisons_admin(data, auteur, date_iso):
    livraisons = data.get("livraisons", [])
    chantier   = data.get("chantier", "Tous chantiers")
    semaine    = data.get("semaine", "")

    if not livraisons:
        from flask import abort
        return jsonify({"erreur": "Aucune livraison"}), 400

    wb = _make_wb()

    # ── Feuille 1 : Récapitulatif ──────────────────────────────────────────────
    cols_recap = ["Matériau","Quantité totale","Unité","Nb livraisons","Prix unitaire","Coût total HT"]
    largeurs_recap = [28, 16, 10, 14, 16, 18]

    ws1 = wb.create_sheet("Récapitulatif")
    _col_widths(ws1, largeurs_recap)
    sous_t = f"{chantier}" + (f" — {semaine}" if semaine else "")
    r = _cartouche(ws1, "Livraisons Chantier — Récapitulatif", sous_t, auteur, len(cols_recap))
    _entetes(ws1, r, cols_recap); r += 1

    # Agréger
    agg = {}
    for l in livraisons:
        el = l.get("element","?")
        if el not in agg:
            agg[el] = {"element":el,"unite":l.get("unite",""),"qte":0,"nb":0,"prix":l.get("prix_unitaire")}
        agg[el]["qte"] += float(l.get("quantite",0) or 0)
        agg[el]["nb"]  += 1
        if l.get("prix_unitaire") and not agg[el]["prix"]:
            agg[el]["prix"] = l.get("prix_unitaire")

    total_cout = 0
    for idx, v in enumerate(sorted(agg.values(), key=lambda x: x["qte"], reverse=True)):
        prix = v.get("prix")
        cout = round(v["qte"] * prix, 2) if prix else None
        if cout: total_cout += cout
        row_data = [
            v["element"], round(v["qte"],3), v["unite"], v["nb"],
            f"{prix} €/{v['unite']}" if prix else "—",
            f"{cout:.2f} €" if cout else "—"
        ]
        _ligne(ws1, r, row_data, idx); r += 1

    # Ligne total
    ws1.row_dimensions[r].height = 22
    for col in range(1, len(cols_recap)+1):
        ws1.cell(row=r, column=col).fill = _fill(_BLEU_ACIER)
    ws1.cell(row=r, column=1, value="TOTAL").font = _font(11, bold=True, color=_BLANC)
    ws1.cell(row=r, column=1).alignment = _align('left','center')
    c = ws1.cell(row=r, column=6, value=f"{total_cout:.2f} €" if total_cout else "—")
    c.font = _font(11, bold=True, color=_BLANC); c.alignment = _align('left','center')

    # ── Feuille 2 : Détail livraisons ─────────────────────────────────────────
    cols_det = ["Date","Chantier","Chef chantier","Matériau","Quantité","Unité","Prix unit.","Coût HT","Notes"]
    largeurs_det = [14, 22, 18, 22, 12, 10, 14, 14, 28]

    ws2 = wb.create_sheet("Détail livraisons")
    _col_widths(ws2, largeurs_det)
    r = _cartouche(ws2, "Livraisons Chantier — Détail", sous_t, auteur, len(cols_det))
    _entetes(ws2, r, cols_det); r += 1

    for idx, l in enumerate(sorted(livraisons, key=lambda x: x.get("date_liv",""), reverse=True)):
        date_fmt = l.get("date_liv","—")
        try:
            from datetime import date
            date_fmt = date.fromisoformat(date_fmt).strftime("%d/%m/%Y")
        except: pass
        prix = l.get("prix_unitaire")
        cout = round(float(l.get("quantite",0) or 0) * prix, 2) if prix else None
        row_data = [
            date_fmt, l.get("chantier","—"), l.get("auteur","—"),
            l.get("element","—"), round(float(l.get("quantite",0) or 0),3),
            l.get("unite","—"), f"{prix} €" if prix else "—",
            f"{cout:.2f} €" if cout else "—", l.get("notes","—") or "—"
        ]
        _ligne(ws2, r, row_data, idx); r += 1

    ws1.freeze_panes = ws2.freeze_panes = "A8"
    return _send_wb(wb, f"FGS-Livraisons-{chantier.replace(' ','-')}-{date_iso}")


# ── Export Livraisons chef ─────────────────────────────────────────────────────
def _export_livraisons_chef(data, auteur, date_iso):
    livraisons = data.get("livraisons", [])
    if not livraisons:
        return jsonify({"erreur": "Aucune livraison"}), 400

    cols = ["Date","Chantier","Matériau","Quantité","Unité","Notes"]
    largeurs = [14, 22, 24, 12, 10, 32]

    wb = _make_wb()
    ws = wb.create_sheet("Mes livraisons")
    _col_widths(ws, largeurs)
    r = _cartouche(ws, "Mes Livraisons Chantier",
                   f"{auteur}  —  {len(livraisons)} livraison(s)", auteur, len(cols))
    _entetes(ws, r, cols); r += 1

    for idx, l in enumerate(sorted(livraisons, key=lambda x: x.get("date_liv",""), reverse=True)):
        date_fmt = l.get("date_liv","—")
        try:
            from datetime import date
            date_fmt = date.fromisoformat(date_fmt).strftime("%d/%m/%Y")
        except: pass
        row_data = [date_fmt, l.get("chantier","—"), l.get("element","—"),
                    round(float(l.get("quantite",0) or 0),3), l.get("unite","—"),
                    l.get("notes","—") or "—"]
        _ligne(ws, r, row_data, idx); r += 1

    ws.freeze_panes = "A8"
    return _send_wb(wb, f"FGS-MesLivraisons-{date_iso}")


# ── Export Total chantier ──────────────────────────────────────────────────────
def _export_total_chantier(data, bd, auteur, date_iso):
    chantier   = data.get("chantier","—")
    pointages  = data.get("pointages",[])
    livraisons = data.get("livraisons",[])
    prix_ref   = data.get("prix_ref",{})   # {element: prix}

    # Calculs MO
    total_h   = sum(float(l.get("heures",0) or 0) for p in pointages for l in (p.get("lignes") or []))
    nb_pres   = sum(len(p.get("lignes") or []) for p in pointages)
    nb_gd     = sum(1 for p in pointages for l in (p.get("lignes") or []) if l.get("gd"))
    nb_pan    = sum(1 for p in pointages for l in (p.get("lignes") or []) if l.get("panier"))
    taux_h    = prix_ref.get("__heures__")
    cout_mo   = round(total_h * taux_h, 2) if taux_h else None

    # Calculs matériaux
    agg_mat = {}
    for l in livraisons:
        el = l.get("element","?")
        if el not in agg_mat: agg_mat[el] = {"qte":0,"unite":l.get("unite","")}
        agg_mat[el]["qte"] += float(l.get("quantite",0) or 0)

    total_mat = 0
    mat_lignes = []
    for el, v in agg_mat.items():
        prix = prix_ref.get(el) or l.get("prix_unitaire") if livraisons else None
        # Chercher prix dans les livraisons
        for lv in livraisons:
            if lv.get("element") == el and lv.get("prix_unitaire"):
                prix = lv.get("prix_unitaire"); break
        cout = round(v["qte"] * prix, 2) if prix else None
        if cout: total_mat += cout
        mat_lignes.append((el, round(v["qte"],3), v["unite"], prix, cout))

    total_gen = (cout_mo or 0) + total_mat
    wb = _make_wb()

    # ── Feuille 1 : Synthèse ──────────────────────────────────────────────────
    cols_s = ["Poste","Détail","Taux","Montant HT"]
    largeurs_s = [22, 45, 16, 18]
    ws1 = wb.create_sheet("Synthèse")
    _col_widths(ws1, largeurs_s)
    r = _cartouche(ws1, f"Synthèse Chantier — {chantier}",
                   "Main d'œuvre & Matériaux", auteur, len(cols_s))
    _entetes(ws1, r, cols_s); r += 1
    _ligne(ws1, r, [
        "Main d'œuvre",
        f"{total_h:.1f} h · {nb_pres} présences · {nb_gd} GD · {nb_pan} paniers",
        f"{taux_h} €/h" if taux_h else "—",
        f"{cout_mo:.2f} €" if cout_mo else "—"
    ], 0); r += 1
    _ligne(ws1, r, [
        "Matériaux", ", ".join(agg_mat.keys()) or "—", "—",
        f"{total_mat:.2f} €" if total_mat else "—"
    ], 1); r += 1
    # Total général
    ws1.row_dimensions[r].height = 24
    for col in range(1, len(cols_s)+1):
        ws1.cell(row=r, column=col).fill = _fill(_ROUGE)
    ws1.cell(row=r, column=1, value="TOTAL GÉNÉRAL").font = _font(12, bold=True, color=_BLANC)
    ws1.cell(row=r, column=1).alignment = _align('left','center')
    c = ws1.cell(row=r, column=4, value=f"{total_gen:.2f} €" if total_gen else "—")
    c.font = _font(12, bold=True, color=_BLANC); c.alignment = _align('left','center')

    # ── Feuille 2 : Matériaux ──────────────────────────────────────────────────
    cols_m = ["Matériau","Quantité","Unité","Prix unitaire","Coût HT"]
    largeurs_m = [28, 14, 10, 16, 16]
    ws2 = wb.create_sheet("Matériaux")
    _col_widths(ws2, largeurs_m)
    r = _cartouche(ws2, "Matériaux Livrés", chantier, auteur, len(cols_m))
    _entetes(ws2, r, cols_m); r += 1
    for idx, (el, qte, unite, prix, cout) in enumerate(mat_lignes):
        _ligne(ws2, r, [el, qte, unite, f"{prix} €" if prix else "—", f"{cout:.2f} €" if cout else "—"], idx); r += 1

    # ── Feuille 3 : Pointages ──────────────────────────────────────────────────
    cols_p = ["Date","Saisi par","Employés","Présences","Total heures","GD","Paniers"]
    largeurs_p = [14, 18, 36, 12, 14, 10, 10]
    ws3 = wb.create_sheet("Pointages")
    _col_widths(ws3, largeurs_p)
    r = _cartouche(ws3, "Pointages — Détail", chantier, auteur, len(cols_p))
    _entetes(ws3, r, cols_p); r += 1
    for idx, ptg in enumerate(sorted(pointages, key=lambda x: x.get("date_jour",""), reverse=True)):
        lignes = ptg.get("lignes") or []
        th = sum(float(l.get("heures",0) or 0) for l in lignes)
        date_fmt = ptg.get("date_jour","—")
        try:
            from datetime import date
            date_fmt = date.fromisoformat(date_fmt).strftime("%d/%m/%Y")
        except: pass
        row_data = [
            date_fmt, ptg.get("auteur","—"),
            ", ".join(l.get("nom","?") for l in lignes) or "—",
            len(lignes), round(th,1),
            sum(1 for l in lignes if l.get("gd")) or "—",
            sum(1 for l in lignes if l.get("panier")) or "—"
        ]
        _ligne(ws3, r, row_data, idx); r += 1

    ws1.freeze_panes = ws2.freeze_panes = ws3.freeze_panes = "A8"
    return _send_wb(wb, f"FGS-Total-{chantier.replace(' ','-')}-{date_iso}")

