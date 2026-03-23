#!/usr/bin/env python3
"""
FGS App — Serveur avec base de données PostgreSQL
Utilise pg8000 (compatible Python 3.14+)
"""

import json, os, re, io, base64
import pg8000.native
from flask import Flask, request, jsonify, send_from_directory, send_file

app = Flask(__name__, static_folder=".")

# ── Connexion ─────────────────────────────────────────────────────────────────

def parse_db_url(url):
    """Parse une DATABASE_URL en paramètres de connexion."""
    if url.startswith("postgres://"):
        url = url.replace("postgres://", "postgresql://", 1)
    m = re.match(r'postgresql://([^:]+):([^@]+)@([^:/]+):?(\d*)/(.+)', url)
    if not m:
        raise ValueError(f"DATABASE_URL invalide : {url}")
    user, password, host, port, dbname = m.groups()
    return {
        "user": user,
        "password": password,
        "host": host,
        "port": int(port) if port else 5432,
        "database": dbname.split("?")[0],  # ignore ?sslmode=...
        "ssl_context": True,               # Render exige SSL
    }

def get_conn():
    url = os.environ.get("DATABASE_URL", "")
    if not url:
        raise RuntimeError("DATABASE_URL manquante.")
    p = parse_db_url(url)
    return pg8000.native.Connection(**p)

# ── Schéma et données par défaut ──────────────────────────────────────────────

SCHEMA_BD = """
CREATE TABLE IF NOT EXISTS fgs_data (
    id     SERIAL PRIMARY KEY,
    cle    TEXT UNIQUE NOT NULL,
    valeur TEXT NOT NULL,
    maj    TIMESTAMPTZ DEFAULT NOW()
)
"""

SCHEMA_POINTAGE = """
CREATE TABLE IF NOT EXISTS pointages (
    id          SERIAL PRIMARY KEY,
    date_jour   DATE NOT NULL,
    chantier    TEXT NOT NULL,
    auteur      TEXT NOT NULL,
    role_auteur TEXT NOT NULL DEFAULT 'chef',
    lignes      JSONB NOT NULL DEFAULT '[]',
    cree_le     TIMESTAMPTZ DEFAULT NOW(),
    maj_le      TIMESTAMPTZ DEFAULT NOW()
)
"""

SCHEMA_LIVCHANTIER = """
CREATE TABLE IF NOT EXISTS livraisons_chantier (
    id           SERIAL PRIMARY KEY,
    date_liv     DATE NOT NULL,
    chantier     TEXT NOT NULL,
    auteur       TEXT NOT NULL,
    element      TEXT NOT NULL,
    quantite     NUMERIC(12,3) NOT NULL DEFAULT 0,
    unite        TEXT NOT NULL DEFAULT 'm3',
    notes        TEXT DEFAULT '',
    cree_le      TIMESTAMPTZ DEFAULT NOW()
)
"""

SCHEMA_PRIX_REF = """
CREATE TABLE IF NOT EXISTS prix_reference (
    id        SERIAL PRIMARY KEY,
    chantier  TEXT NOT NULL,
    element   TEXT NOT NULL,
    prix_unitaire NUMERIC(12,2) NOT NULL,
    maj_le    TIMESTAMPTZ DEFAULT NOW(),
    UNIQUE(chantier, element)
)
"""

DONNEES_DEFAUT = {
    "categories": [
        {"id":"c1","nom":"Pelles","icone":"🦾"},
        {"id":"c2","nom":"Projection voie sèche","icone":"💨"},
        {"id":"c3","nom":"Injection","icone":"💉"},
        {"id":"c4","nom":"Manutention","icone":"🏗️"},
        {"id":"c5","nom":"Foreuses","icone":"🔩"},
        {"id":"c6","nom":"Glissières","icone":"🛤️"},
        {"id":"c7","nom":"Disqueuses thermiques","icone":"⚙️"},
    ],
    "machines": [
        {"id":"EQ-001","catId":"c1","nom":"Minipelle 2,5t","modele":"Kubota U25-3",
         "annee":2021,"heures":1620,"hEntretien":1500,"seuil":250,"vgp":"2026-12-03",
         "site":"Lyon","serie":"KB-U25-3-XYZ","poids":2500,
         "piecesAssociees":[
             {"pieceId":"P-001","heuresInstallation":1500,"dateInstallation":"2026-01-10"},
             {"pieceId":"P-003","heuresInstallation":1200,"dateInstallation":"2025-10-01"},
         ]},
        {"id":"EQ-002","catId":"c4","nom":"Chariot télescopique","modele":"Manitou MT625",
         "annee":2020,"heures":3100,"hEntretien":2900,"seuil":200,"vgp":"2026-04-15",
         "site":"Grenoble","piecesAssociees":[]},
    ],
    "pieces": [
        {"id":"P-001","nom":"Filtre huile moteur","ref":"FH-123","dureeVal":250,"dureeUnite":"heures","stock":3,"notes":"Utiliser pièce d'origine"},
        {"id":"P-002","nom":"Courroie de distribution","ref":"CR-321","dureeVal":500,"dureeUnite":"heures","stock":1},
        {"id":"P-003","nom":"Filtre à air","ref":"FA-456","dureeVal":500,"dureeUnite":"heures","stock":2},
        {"id":"P-004","nom":"Filtre hydraulique","ref":"FH-789","dureeVal":1000,"dureeUnite":"heures","stock":0},
    ],
    "interventions": [
        {"id":"I-001","machineId":"EQ-001","date":"2026-01-10","datePrevue":"","heures":1500,
         "type":"entretien","notes":"Vidange + filtre huile","piecesChangees":["P-001"]},
        {"id":"I-002","machineId":"EQ-002","date":"2026-02-01","datePrevue":"","heures":2900,
         "type":"entretien","notes":"Révision à 2900h","piecesChangees":[]},
    ],
    "livraisons": [],
}

SCHEMA_CHANTIERS = """
CREATE TABLE IF NOT EXISTS chantiers (
    id           SERIAL PRIMARY KEY,
    nom          TEXT NOT NULL,
    localisation TEXT NOT NULL DEFAULT '',
    lat          DOUBLE PRECISION,
    lng          DOUBLE PRECISION,
    date_debut   DATE,
    date_fin     DATE,
    cree_par     TEXT NOT NULL DEFAULT '',
    cree_le      TIMESTAMPTZ DEFAULT NOW(),
    maj_le       TIMESTAMPTZ DEFAULT NOW()
)
"""

SCHEMA_PERSONNEL = """
CREATE TABLE IF NOT EXISTS personnel (
    id            SERIAL PRIMARY KEY,
    nom           TEXT NOT NULL,
    prenom        TEXT NOT NULL DEFAULT '',
    poste         TEXT NOT NULL DEFAULT '',
    type_contrat  TEXT NOT NULL DEFAULT 'CDI',
    cout_horaire  NUMERIC(10,2),
    adresse       TEXT DEFAULT '',
    urgence_nom   TEXT DEFAULT '',
    urgence_tel   TEXT DEFAULT '',
    notes         TEXT DEFAULT '',
    cree_par      TEXT DEFAULT '',
    cree_le       TIMESTAMPTZ DEFAULT NOW(),
    maj_le        TIMESTAMPTZ DEFAULT NOW()
)
"""

SCHEMA_PERSONNEL_MIGRATION = """
ALTER TABLE personnel ADD COLUMN IF NOT EXISTS type_contrat TEXT NOT NULL DEFAULT 'CDI'
"""

def init_db():
    conn = get_conn()
    conn.run(SCHEMA_BD)
    conn.run(SCHEMA_POINTAGE)
    conn.run(SCHEMA_LIVCHANTIER)
    conn.run(SCHEMA_PRIX_REF)
    conn.run(SCHEMA_CHANTIERS)
    conn.run(SCHEMA_PERSONNEL)
    try:
        conn.run(SCHEMA_PERSONNEL_MIGRATION)
    except Exception:
        pass  # colonne déjà présente
    rows = conn.run("SELECT COUNT(*) FROM fgs_data WHERE cle = 'bd'")
    if rows[0][0] == 0:
        conn.run(
            "INSERT INTO fgs_data (cle, valeur) VALUES (:cle, :valeur)",
            cle="bd",
            valeur=json.dumps(DONNEES_DEFAUT, ensure_ascii=False)
        )
    conn.close()
    print("Base PostgreSQL initialisee (fgs_data + pointages + chantiers + personnel).")

# ── Lecture / écriture ────────────────────────────────────────────────────────

def lire():
    conn = get_conn()
    rows = conn.run("SELECT valeur FROM fgs_data WHERE cle = 'bd'")
    conn.close()
    if not rows:
        init_db()
        return DONNEES_DEFAUT
    val = rows[0][0]
    return json.loads(val) if isinstance(val, str) else val

def ecrire(bd):
    conn = get_conn()
    conn.run(
        """
        INSERT INTO fgs_data (cle, valeur, maj) VALUES (:cle, :valeur, NOW())
        ON CONFLICT (cle) DO UPDATE SET valeur = EXCLUDED.valeur, maj = NOW()
        """,
        cle="bd",
        valeur=json.dumps(bd, ensure_ascii=False)
    )
    conn.close()

# ── Routes ────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return send_from_directory(".", "app.html")


def get_role(identifiant):
    """
    Retourne le rôle d'un utilisateur depuis ses variables d'environnement.
    Format de la variable : identifiant=motdepasse:role
    Rôles : admin, rh, chef
    Exemple : Admin=MonMDP:admin  /  rh.martin=MDP:rh  /  chef.lyon=MDP:chef
    Compatibilité ancien format (juste mot de passe sans rôle) = chef par défaut
    """
    for key, value in os.environ.items():
        if key.lower() == identifiant.lower():
            parts = value.split(":")
            if len(parts) >= 2:
                return parts[-1].strip().lower()  # dernier segment = rôle
            return "chef"  # ancien format sans rôle = chef par défaut
    return "chef"


@app.route("/api/whoami")
def whoami():
    """Debug: retourne les variables d'env disponibles (sans les valeurs) pour diagnostic"""
    comptes = []
    for key, value in os.environ.items():
        if key.lower() in ('admin', 'rh') or (len(key) < 20 and not key.startswith('_') and not key.startswith('RENDER') and not key.startswith('DATABASE') and not key.startswith('PORT') and not key.startswith('PATH') and not key.startswith('HOME') and not key.startswith('USER') and not key.startswith('PWD') and not key.startswith('LANG') and not key.startswith('LC_') and not key.startswith('PYTHON') and not key.startswith('PIP') and not key.startswith('VIRTUAL')):
            # Déduire le rôle comme auth() le ferait
            id_lower = key.lower()
            role_deduit = "admin" if id_lower == "admin" else ("rh" if id_lower == "rh" else "chef")
            has_colon = ":" in value
            if has_colon:
                last_colon = value.rfind(":")
                role_candidat = value[last_colon+1:].strip().lower()
                if role_candidat in ("chef", "admin", "rh"):
                    role_deduit = role_candidat
            comptes.append({"identifiant": key, "role_deduit": role_deduit, "format": "nouveau" if has_colon else "ancien"})
    return jsonify({"comptes": comptes})

@app.route("/api/auth", methods=["POST"])
def auth():
    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"ok": False}), 400
        identifiant = data.get("identifiant", "").strip()
        mdp = data.get("mdp", "")
        if not identifiant or not mdp:
            return jsonify({"ok": False}), 400
        for key, value in os.environ.items():
            if key.lower() != identifiant.lower():
                continue
            value = value.strip()
            # Format variable Render : "motdepasse:role" (ex: MonMDP:admin, MonMDP:chef)
            # ou "motdepasse" seul (ancien format sans rôle -> chef par défaut)
            # L'utilisateur saisit UNIQUEMENT le mot de passe, sans le :role
            role = "chef"
            mdp_stocke = value
            if ":" in value:
                last_colon = value.rfind(":")
                role_candidat = value[last_colon+1:].strip().lower()
                if role_candidat in ("chef", "admin", "rh"):
                    role = role_candidat
                    mdp_stocke = value[:last_colon]
                # sinon le ":" fait partie du mdp (ex: "Illite@8020") -> mdp_stocke = value entier
            if mdp_stocke == mdp:
                return jsonify({"ok": True, "nom": key, "role": role})
        return jsonify({"ok": False})
    except Exception as e:
        return jsonify({"ok": False, "erreur": str(e)}), 500



@app.route("/api/pointages", methods=["GET"])
def get_pointages():
    """Retourne les pointages selon le rôle."""
    try:
        role   = request.args.get("role", "chef")
        auteur = request.args.get("auteur", "")
        date_debut = request.args.get("debut", "")
        date_fin   = request.args.get("fin", "")

        conn = get_conn()
        if role in ("admin", "rh"):
            if date_debut and date_fin:
                rows = conn.run(
                    "SELECT * FROM pointages WHERE date_jour BETWEEN :d AND :f ORDER BY date_jour DESC, cree_le DESC",
                    d=date_debut, f=date_fin
                )
            else:
                rows = conn.run("SELECT * FROM pointages ORDER BY date_jour DESC, cree_le DESC LIMIT 200")
        else:
            # Chef : uniquement ses propres pointages
            if date_debut and date_fin:
                rows = conn.run(
                    "SELECT * FROM pointages WHERE auteur ILIKE :a AND date_jour BETWEEN :d AND :f ORDER BY date_jour DESC",
                    a=auteur, d=date_debut, f=date_fin
                )
            else:
                rows = conn.run(
                    "SELECT * FROM pointages WHERE auteur ILIKE :a ORDER BY date_jour DESC LIMIT 100",
                    a=auteur
                )
        conn.close()

        cols = ["id","date_jour","chantier","auteur","role_auteur","lignes","cree_le","maj_le"]
        result = []
        for row in rows:
            d = dict(zip(cols, row))
            d["date_jour"] = str(d["date_jour"])
            d["cree_le"]   = str(d["cree_le"])
            d["maj_le"]    = str(d["maj_le"])
            if isinstance(d["lignes"], str):
                d["lignes"] = json.loads(d["lignes"])
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/pointages", methods=["POST"])
def save_pointage():
    """Crée ou met à jour un pointage."""
    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"erreur": "Données invalides"}), 400

        pid        = data.get("id")
        date_jour  = data.get("date_jour")
        chantier   = data.get("chantier", "")
        auteur     = data.get("auteur", "")
        role_auteur= data.get("role_auteur", "chef")
        lignes     = data.get("lignes", [])

        conn = get_conn()
        if pid:
            conn.run(
                "UPDATE pointages SET lignes=:l, chantier=:c, maj_le=NOW() WHERE id=:id",
                l=json.dumps(lignes, ensure_ascii=False), c=chantier, id=pid
            )
            new_id = pid
        else:
            rows = conn.run(
                """INSERT INTO pointages (date_jour, chantier, auteur, role_auteur, lignes)
                   VALUES (:d, :c, :a, :r, :l) RETURNING id""",
                d=date_jour, c=chantier, a=auteur, r=role_auteur,
                l=json.dumps(lignes, ensure_ascii=False)
            )
            new_id = rows[0][0]
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/pointages/<int:pid>", methods=["DELETE"])
def delete_pointage(pid):
    try:
        role = request.args.get("role", "chef")
        auteur = request.args.get("auteur", "")
        conn = get_conn()
        if role in ("admin", "rh"):
            conn.run("DELETE FROM pointages WHERE id=:id", id=pid)
        else:
            conn.run("DELETE FROM pointages WHERE id=:id AND auteur ILIKE :a", id=pid, a=auteur)
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500




@app.route("/api/bd", methods=["GET"])
def get_bd():
    try:
        return jsonify(lire())
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500

@app.route("/api/bd", methods=["POST"])
def post_bd():
    try:
        bd = request.get_json(force=True, silent=True)
        if bd is None:
            return jsonify({"erreur": "JSON invalide"}), 400
        ecrire(bd)
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/backup")
def backup():
    """
    Télécharge une sauvegarde complète des données au format JSON.
    Protégé par une clé secrète : /api/backup?key=VOTRE_CLE
    
    Sur Render : ajoutez une variable d'environnement BACKUP_KEY = votre_mot_de_passe_secret
    """
    cle = request.args.get("key", "")
    cle_attendue = os.environ.get("BACKUP_KEY", "")
    
    if not cle_attendue:
        return jsonify({"erreur": "Variable BACKUP_KEY non configurée sur Render."}), 500
    if cle != cle_attendue:
        return jsonify({"erreur": "Clé incorrecte."}), 403
    
    try:
        bd = lire()
        from datetime import datetime
        date_str = datetime.now().strftime("%Y-%m-%d_%H-%M")
        nom_fichier = f"fgs-backup-{date_str}.json"
        contenu = json.dumps(bd, ensure_ascii=False, indent=2)
        
        from flask import Response
        return Response(
            contenu,
            mimetype="application/json",
            headers={"Content-Disposition": f"attachment; filename={nom_fichier}"}
        )
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/restore", methods=["POST"])
def restore():
    """
    Restaure les données depuis un fichier JSON.
    Protégé par la même clé secrète BACKUP_KEY.
    
    Utilisation : POST /api/restore?key=VOTRE_CLE
    Corps : le fichier JSON de sauvegarde
    """
    cle = request.args.get("key", "")
    cle_attendue = os.environ.get("BACKUP_KEY", "")
    
    if not cle_attendue:
        return jsonify({"erreur": "Variable BACKUP_KEY non configurée sur Render."}), 500
    if cle != cle_attendue:
        return jsonify({"erreur": "Clé incorrecte."}), 403
    
    try:
        bd = request.get_json(force=True, silent=True)
        if not bd:
            return jsonify({"erreur": "Fichier JSON invalide ou vide."}), 400
        
        # Vérification basique que c'est bien une sauvegarde FGS
        champs_requis = ["categories", "machines", "pieces", "interventions", "livraisons"]
        for champ in champs_requis:
            if champ not in bd:
                return jsonify({"erreur": f"Sauvegarde invalide : champ '{champ}' manquant."}), 400
        
        ecrire(bd)
        return jsonify({
            "ok": True,
            "message": f"Restauration réussie : {len(bd.get('machines',[]))} machines, {len(bd.get('pieces',[]))} pièces, {len(bd.get('interventions',[]))} interventions."
        })
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/livchantier", methods=["GET"])
def get_livchantier():
    try:
        role    = request.args.get("role", "chef")
        auteur  = request.args.get("auteur", "")
        chantier= request.args.get("chantier", "")
        element = request.args.get("element", "")
        chef_f  = request.args.get("chef", "")
        debut   = request.args.get("debut", "")
        fin     = request.args.get("fin", "")

        where = []
        params = {}
        if role not in ("admin", "rh"):
            where.append("auteur ILIKE :auteur")
            params["auteur"] = auteur
        if chantier:
            where.append("chantier ILIKE :chantier")
            params["chantier"] = f"%{chantier}%"
        if element:
            where.append("element = :element")
            params["element"] = element
        if chef_f:
            where.append("auteur ILIKE :chef_f")
            params["chef_f"] = f"%{chef_f}%"
        if debut and fin:
            where.append("date_liv BETWEEN :debut AND :fin")
            params["debut"] = debut
            params["fin"] = fin

        sql = """SELECT l.id, l.date_liv, l.chantier, l.auteur, l.element,
                        l.quantite, l.unite, p.prix_unitaire, l.notes, l.cree_le
                 FROM livraisons_chantier l
                 LEFT JOIN prix_reference p ON p.chantier=l.chantier AND p.element=l.element"""
        if where:
            sql += " WHERE " + " AND ".join(where)
        sql += " ORDER BY date_liv DESC, cree_le DESC"

        conn = get_conn()
        rows = conn.run(sql, **params) if params else conn.run(sql)
        conn.close()

        cols = ["id","date_liv","chantier","auteur","element","quantite","unite","prix_unitaire","notes","cree_le"]
        result = []
        for row in rows:
            d = dict(zip(cols, row))
            d["date_liv"] = str(d["date_liv"])
            d["cree_le"]  = str(d["cree_le"])
            d["quantite"] = float(d["quantite"]) if d["quantite"] is not None else 0
            d["prix_unitaire"] = float(d["prix_unitaire"]) if d["prix_unitaire"] is not None else None
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/livchantier", methods=["POST"])
def save_livchantier():
    try:
        data = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"erreur": "Données invalides"}), 400
        lid      = data.get("id")
        date_liv = data.get("date_liv")
        chantier = data.get("chantier", "")
        auteur   = data.get("auteur", "")
        element  = data.get("element", "béton")
        quantite = float(data.get("quantite", 0))
        unite    = data.get("unite", "m3")
        notes    = data.get("notes", "")
        conn = get_conn()
        if lid:
            conn.run(
                """UPDATE livraisons_chantier
                   SET date_liv=:d, chantier=:c, element=:e, quantite=:q, unite=:u, notes=:n
                   WHERE id=:id""",
                d=date_liv, c=chantier, e=element, q=quantite, u=unite, n=notes, id=lid
            )
            new_id = lid
        else:
            rows = conn.run(
                """INSERT INTO livraisons_chantier
                   (date_liv,chantier,auteur,element,quantite,unite,notes)
                   VALUES (:d,:c,:a,:e,:q,:u,:n) RETURNING id""",
                d=date_liv, c=chantier, a=auteur, e=element, q=quantite, u=unite, n=notes
            )
            new_id = rows[0][0]
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/livchantier/<int:lid>", methods=["DELETE"])
def delete_livchantier(lid):
    try:
        role   = request.args.get("role", "chef")
        auteur = request.args.get("auteur", "")
        conn = get_conn()
        if role in ("admin", "rh"):
            conn.run("DELETE FROM livraisons_chantier WHERE id=:id", id=lid)
        else:
            conn.run("DELETE FROM livraisons_chantier WHERE id=:id AND auteur ILIKE :a",
                     id=lid, a=auteur)
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/prix_reference", methods=["GET"])
def get_prix_ref():
    """Retourne tous les prix de référence (chantier x element)."""
    try:
        conn = get_conn()
        rows = conn.run("SELECT id, chantier, element, prix_unitaire, maj_le FROM prix_reference ORDER BY chantier, element")
        conn.close()
        cols = ["id","chantier","element","prix_unitaire","maj_le"]
        result = []
        for row in rows:
            d = dict(zip(cols, row))
            d["prix_unitaire"] = float(d["prix_unitaire"]) if d["prix_unitaire"] else None
            d["maj_le"] = str(d["maj_le"])
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/prix_reference", methods=["POST"])
def save_prix_ref():
    """Upsert un prix de référence pour un couple chantier+element."""
    try:
        data = request.get_json(force=True, silent=True)
        chantier = data.get("chantier", "").strip()
        element  = data.get("element", "").strip()
        prix     = float(data.get("prix_unitaire", 0))
        if not chantier or not element:
            return jsonify({"erreur": "chantier et element requis"}), 400
        conn = get_conn()
        conn.run(
            """INSERT INTO prix_reference (chantier, element, prix_unitaire, maj_le)
               VALUES (:c, :e, :p, NOW())
               ON CONFLICT (chantier, element)
               DO UPDATE SET prix_unitaire=EXCLUDED.prix_unitaire, maj_le=NOW()""",
            c=chantier, e=element, p=prix
        )
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500

@app.route("/api/sante")
def sante():
    try:
        conn = get_conn()
        rows = conn.run("SELECT maj FROM fgs_data WHERE cle='bd'")
        conn.close()
        derniere_maj = str(rows[0][0]) if rows else "jamais"
        return jsonify({"ok": True, "derniere_maj": derniere_maj})
    except Exception as e:
        return jsonify({"ok": False, "erreur": str(e)}), 500


# ── Chantiers ─────────────────────────────────────────────────────────────────

@app.route("/api/chantiers", methods=["GET"])
def get_chantiers():
    try:
        conn = get_conn()
        rows = conn.run(
            "SELECT id, nom, localisation, lat, lng, date_debut, date_fin, cree_par, cree_le "
            "FROM chantiers ORDER BY cree_le DESC"
        )
        conn.close()
        cols = ["id", "nom", "localisation", "lat", "lng", "date_debut", "date_fin", "cree_par", "cree_le"]
        result = []
        for row in rows:
            d = dict(zip(cols, row))
            d["date_debut"] = str(d["date_debut"]) if d["date_debut"] else None
            d["date_fin"]   = str(d["date_fin"])   if d["date_fin"]   else None
            d["cree_le"]    = str(d["cree_le"])
            d["lat"]        = float(d["lat"])  if d["lat"]  is not None else None
            d["lng"]        = float(d["lng"])  if d["lng"]  is not None else None
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/chantiers", methods=["POST"])
def save_chantier():
    try:
        data         = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"erreur": "Données invalides"}), 400
        cid          = data.get("id")
        nom          = data.get("nom", "").strip()
        localisation = data.get("localisation", "").strip()
        lat          = data.get("lat")
        lng          = data.get("lng")
        date_debut   = data.get("date_debut") or None
        date_fin     = data.get("date_fin")   or None
        cree_par     = data.get("cree_par", "")
        if not nom:
            return jsonify({"erreur": "Nom obligatoire"}), 400
        lat = float(lat) if lat is not None else None
        lng = float(lng) if lng is not None else None
        conn = get_conn()
        if cid:
            conn.run(
                """UPDATE chantiers SET nom=:nom, localisation=:loc, lat=:lat, lng=:lng,
                   date_debut=:dd, date_fin=:df, maj_le=NOW() WHERE id=:id""",
                nom=nom, loc=localisation, lat=lat, lng=lng,
                dd=date_debut, df=date_fin, id=cid
            )
            new_id = cid
        else:
            rows = conn.run(
                """INSERT INTO chantiers (nom, localisation, lat, lng, date_debut, date_fin, cree_par)
                   VALUES (:nom, :loc, :lat, :lng, :dd, :df, :cp) RETURNING id""",
                nom=nom, loc=localisation, lat=lat, lng=lng,
                dd=date_debut, df=date_fin, cp=cree_par
            )
            new_id = rows[0][0]
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/chantiers/<int:cid>", methods=["DELETE"])
def delete_chantier(cid):
    try:
        conn = get_conn()
        conn.run("DELETE FROM chantiers WHERE id=:id", id=cid)
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


# ── Personnel ─────────────────────────────────────────────────────────────────

@app.route("/api/personnel", methods=["GET"])
def get_personnel():
    try:
        role = request.args.get("role", "chef")
        conn = get_conn()
        rows = conn.run(
            "SELECT id, nom, prenom, poste, type_contrat, cout_horaire, adresse, urgence_nom, urgence_tel, notes, cree_par, cree_le "
            "FROM personnel ORDER BY nom, prenom"
        )
        conn.close()
        cols = ["id", "nom", "prenom", "poste", "type_contrat", "cout_horaire", "adresse",
                "urgence_nom", "urgence_tel", "notes", "cree_par", "cree_le"]
        result = []
        for row in rows:
            d = dict(zip(cols, row))
            d["cree_le"] = str(d["cree_le"])
            if d["cout_horaire"] is not None:
                if role in ("admin", "rh"):
                    d["cout_horaire"] = float(d["cout_horaire"])
                else:
                    d["cout_horaire"] = None  # masqué aux non-admin
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/personnel", methods=["POST"])
def save_personnel():
    try:
        data        = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"erreur": "Données invalides"}), 400
        pid         = data.get("id")
        nom         = data.get("nom", "").strip()
        prenom      = data.get("prenom", "").strip()
        poste       = data.get("poste", "").strip()
        type_contrat = data.get("type_contrat", "CDI").strip()
        cout_h      = data.get("cout_horaire")
        adresse     = data.get("adresse", "").strip()
        urgence_nom = data.get("urgence_nom", "").strip()
        urgence_tel = data.get("urgence_tel", "").strip()
        notes       = data.get("notes", "").strip()
        cree_par    = data.get("cree_par", "")
        role        = data.get("role", "chef")
        if not nom:
            return jsonify({"erreur": "Nom obligatoire"}), 400
        if type_contrat not in ("CDI", "CDD", "Intérimaire"):
            type_contrat = "CDI"
        cout_h = float(cout_h) if (cout_h is not None and role in ("admin", "rh")) else None
        conn = get_conn()
        if pid:
            if role in ("admin", "rh"):
                conn.run(
                    """UPDATE personnel SET nom=:nom, prenom=:prenom, poste=:poste, type_contrat=:tc, cout_horaire=:ch,
                       adresse=:adr, urgence_nom=:un, urgence_tel=:ut, notes=:no, maj_le=NOW() WHERE id=:id""",
                    nom=nom, prenom=prenom, poste=poste, tc=type_contrat, ch=cout_h,
                    adr=adresse, un=urgence_nom, ut=urgence_tel, no=notes, id=pid
                )
            else:
                conn.run(
                    """UPDATE personnel SET nom=:nom, prenom=:prenom, poste=:poste, type_contrat=:tc,
                       adresse=:adr, urgence_nom=:un, urgence_tel=:ut, notes=:no, maj_le=NOW() WHERE id=:id""",
                    nom=nom, prenom=prenom, poste=poste, tc=type_contrat,
                    adr=adresse, un=urgence_nom, ut=urgence_tel, no=notes, id=pid
                )
            new_id = pid
        else:
            rows = conn.run(
                """INSERT INTO personnel (nom, prenom, poste, type_contrat, cout_horaire, adresse, urgence_nom, urgence_tel, notes, cree_par)
                   VALUES (:nom, :prenom, :poste, :tc, :ch, :adr, :un, :ut, :no, :cp) RETURNING id""",
                nom=nom, prenom=prenom, poste=poste, tc=type_contrat, ch=cout_h,
                adr=adresse, un=urgence_nom, ut=urgence_tel, no=notes, cp=cree_par
            )
            new_id = rows[0][0]
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/personnel/<int:pid>", methods=["DELETE"])
def delete_personnel_route(pid):
    try:
        role = request.args.get("role", "chef")
        if role not in ("admin", "rh"):
            return jsonify({"erreur": "Accès refusé"}), 403
        conn = get_conn()
        conn.run("DELETE FROM personnel WHERE id=:id", id=pid)
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500

# ── Démarrage ─────────────────────────────────────────────────────────────────

# ── Initialisation au démarrage (fonctionne avec Gunicorn ET python direct) ──
# Appelé à l'import du module — crée la table si elle n'existe pas
try:
    init_db()
except Exception as e:
    print(f"AVERTISSEMENT init_db: {e}")

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print(f"Serveur demarre sur http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", port=port, debug=False)


# ── Anomalies ─────────────────────────────────────────────────────────────────

SCHEMA_ANOMALIES = """
CREATE TABLE IF NOT EXISTS anomalies (
    id          SERIAL PRIMARY KEY,
    date_ano    DATE NOT NULL DEFAULT CURRENT_DATE,
    nom_machine TEXT NOT NULL,
    num_parc    TEXT NOT NULL DEFAULT '',
    heures      INTEGER,
    auteur      TEXT NOT NULL,
    description TEXT NOT NULL,
    statut      TEXT NOT NULL DEFAULT 'ouvert',
    cree_le     TIMESTAMPTZ DEFAULT NOW()
)
"""

SCHEMA_ANOMALIES_MIGRATION = """
ALTER TABLE anomalies ADD COLUMN IF NOT EXISTS heures INTEGER
"""

def init_anomalies():
    conn = get_conn()
    conn.run(SCHEMA_ANOMALIES)
    try:
        conn.run(SCHEMA_ANOMALIES_MIGRATION)
    except Exception:
        pass  # colonne déjà existante
    conn.close()

try:
    init_anomalies()
except Exception as e:
    print(f"AVERTISSEMENT init_anomalies: {e}")


@app.route("/api/anomalies", methods=["GET"])
def get_anomalies():
    try:
        role   = request.args.get("role", "chef")
        auteur = request.args.get("auteur", "")
        conn   = get_conn()
        if role in ("admin", "rh"):
            rows = conn.run(
                "SELECT id,date_ano,nom_machine,num_parc,heures,auteur,description,statut,cree_le "
                "FROM anomalies ORDER BY cree_le DESC LIMIT 200"
            )
        else:
            rows = conn.run(
                "SELECT id,date_ano,nom_machine,num_parc,heures,auteur,description,statut,cree_le "
                "FROM anomalies WHERE auteur ILIKE :a ORDER BY cree_le DESC LIMIT 100",
                a=auteur
            )
        conn.close()
        cols = ["id","date_ano","nom_machine","num_parc","heures","auteur","description","statut","cree_le"]
        result = []
        for row in rows:
            d = dict(zip(cols, row))
            d["date_ano"] = str(d["date_ano"])
            d["cree_le"]  = str(d["cree_le"])
            result.append(d)
        return jsonify(result)
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/anomalies", methods=["POST"])
def save_anomalie():
    try:
        data        = request.get_json(force=True, silent=True)
        if not data:
            return jsonify({"erreur": "Données invalides"}), 400
        aid         = data.get("id")
        date_ano    = data.get("date_ano", "")
        nom_machine = data.get("nom_machine", "").strip()
        num_parc    = data.get("num_parc", "").strip()
        heures_val  = data.get("heures")
        heures_val  = int(heures_val) if heures_val else None
        auteur      = data.get("auteur", "")
        description = data.get("description", "").strip()
        statut      = data.get("statut", "ouvert")
        if not nom_machine or not description:
            return jsonify({"erreur": "Champs obligatoires manquants"}), 400
        conn = get_conn()
        if aid:
            conn.run(
                "UPDATE anomalies SET statut=:s, description=:d WHERE id=:id",
                s=statut, d=description, id=aid
            )
            new_id = aid
        else:
            rows = conn.run(
                """INSERT INTO anomalies (date_ano,nom_machine,num_parc,heures,auteur,description,statut)
                   VALUES (:da,:nm,:np,:he,:au,:de,:st) RETURNING id""",
                da=date_ano, nm=nom_machine, np=num_parc, he=heures_val,
                au=auteur, de=description, st=statut
            )
            new_id = rows[0][0]
        conn.close()
        return jsonify({"ok": True, "id": new_id})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/anomalies/<int:aid>", methods=["DELETE"])
def delete_anomalie(aid):
    try:
        role   = request.args.get("role", "chef")
        auteur = request.args.get("auteur", "")
        conn   = get_conn()
        if role in ("admin", "rh"):
            conn.run("DELETE FROM anomalies WHERE id=:id", id=aid)
        else:
            conn.run("DELETE FROM anomalies WHERE id=:id AND auteur ILIKE :a", id=aid, a=auteur)
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


@app.route("/api/anomalies/<int:aid>/statut", methods=["POST"])
def update_statut_anomalie(aid):
    """Admin peut changer le statut : ouvert -> en_cours -> resolu"""
    try:
        data   = request.get_json(force=True, silent=True)
        statut = data.get("statut", "ouvert")
        conn   = get_conn()
        conn.run("UPDATE anomalies SET statut=:s WHERE id=:id", s=statut, id=aid)
        conn.close()
        return jsonify({"ok": True})
    except Exception as e:
        return jsonify({"erreur": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# MOTEUR D'EXPORT XLSX — FGS Travaux Spéciaux
# Optimisé impression A4 — style sobre (cartouche discret, police 11pt)
# Portrait : max 90 units | Paysage : max 145 units (marges 0.47")
# ══════════════════════════════════════════════════════════════════════════════

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

# ── Couleurs (discrètes, sans fonds sombres) ──────────────────────────────────
C_ROUGE    = "C0392B"   # Rouge FGS
C_ROUGE_L  = "FDECEA"   # Rouge très clair  (alerte)
C_ORANGE_L = "FFF3E0"   # Orange très clair (avertissement)
C_VERT_L   = "F1F8F1"   # Vert très clair   (OK)
C_BLEU_L   = "EEF2F7"   # Bleu très clair   (VGP)
C_GRIS_ENT = "EFEFEF"   # Gris en-têtes
C_GRIS_LIG = "FAFAFA"   # Gris ligne paire
C_BORDURE  = "CCCCCC"   # Bordure fine
C_BORD_EP  = "888888"   # Bordure épaisse (séparateurs)
C_TEXTE    = "222222"   # Texte principal
C_BLANC    = "FFFFFF"


# ── Helpers de style ──────────────────────────────────────────────────────────
def _S(c=C_BORDURE, s="thin"):
    return Side(style=s, color=c)

def _fill(c):
    return PatternFill(fill_type="solid", fgColor=c)

def _font(sz=11, bold=False, color=C_TEXTE, italic=False):
    return Font(name="Calibri", size=sz, bold=bold, color=color, italic=italic)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def _set_cols(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _date_fr(d):
    if not d or str(d) in ("—", "None", ""):
        return "—"
    try:
        from datetime import date
        return date.fromisoformat(str(d)).strftime("%d/%m/%Y")
    except:
        return str(d)


# ── Configuration impression A4 ───────────────────────────────────────────────
def _setup_print(ws, orientation="portrait"):
    """
    Configure la feuille pour impression A4 :
    - Même marges que les fichiers FGS de référence (0.47")
    - fitToWidth=1 pour forcer 1 page de large (Excel réduit si besoin)
    - Numéro de page en bas à droite
    - En-têtes répétés sur chaque page
    """
    # Activer fitToPage (propriété sur la feuille)
    if not ws.sheet_properties.pageSetUpPr:
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    else:
        ws.sheet_properties.pageSetUpPr.fitToPage = True

    ws.page_setup.paperSize    = 9             # A4
    ws.page_setup.orientation  = orientation
    ws.page_setup.fitToWidth   = 1             # 1 page de large max
    ws.page_setup.fitToHeight  = 0             # autant de pages en hauteur que nécessaire

    # Marges identiques aux fichiers FGS de référence
    ws.page_margins.left   = 0.4724   # 12 mm
    ws.page_margins.right  = 0.4724
    ws.page_margins.top    = 0.7874   # 20 mm (espace pour cartouche)
    ws.page_margins.bottom = 0.5906   # 15 mm
    ws.page_margins.header = 0.2362   # 6 mm
    ws.page_margins.footer = 0.2362

    # Footer : numéro de page en bas à droite
    ws.oddFooter.right.text  = "Page &P / &N"
    ws.oddFooter.right.size  = 8
    ws.oddFooter.right.color = "999999"

    # Répéter la ligne d'en-têtes (ligne 7) sur chaque page imprimée
    ws.print_title_rows = "7:7"


# ── Cartouche sobre (inspiré Fiche métré FGS) ────────────────────────────────
def _cartouche(ws, titre_doc, auteur, nb_cols):
    """
    Cartouche sur 6 lignes :
      L1-L4 : Logo à gauche (A1), infos document à droite (col C)
      L5     : Espace vide
      L6     : Trait rouge medium (séparateur)
    Retourne 7 (première ligne disponible = en-têtes).
    """
    from datetime import datetime
    now = datetime.now()

    # Hauteurs lignes cartouche (identiques aux fichiers de référence)
    for row, h in [(1, 17), (2, 17), (3, 17), (4, 17), (5, 6), (6, 4)]:
        ws.row_dimensions[row].height = h

    # ── Logo (ancré en A1) ────────────────────────────────────────────────────
    logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logo_fgs.png")
    if os.path.exists(logo_path):
        try:
            img = XLImage(logo_path)
            ratio = img.width / max(img.height, 1)
            img.height = 64
            img.width  = int(64 * ratio)
            img.anchor = "A1"
            ws.add_image(img)
        except Exception:
            c = ws.cell(row=1, column=1, value="FGS")
            c.font = _font(sz=18, bold=True, color=C_ROUGE)

    # ── Infos document (col C, lignes 1-4) ───────────────────────────────────
    infos = [
        (titre_doc,                                      _font(sz=13, bold=True)),
        ("FGS Travaux Spéciaux",                         _font(sz=10, italic=True, color="666666")),
        (f"Émetteur : {auteur}",                         _font(sz=10)),
        (f"Exporté le {now.strftime('%d/%m/%Y')} à {now.strftime('%Hh%M')}",
                                                          _font(sz=9, italic=True, color="999999")),
    ]
    for i, (val, fnt) in enumerate(infos, 1):
        c = ws.cell(row=i, column=3, value=val)
        c.font      = fnt
        c.alignment = _align("left", "center")

    # ── Trait séparateur rouge L6 ─────────────────────────────────────────────
    for col in range(1, nb_cols + 1):
        ws.cell(row=6, column=col).border = Border(
            bottom=Side(style="medium", color=C_ROUGE)
        )

    # Geler le cartouche (lignes 1-7) à l'écran
    ws.freeze_panes = "A8"
    return 7  # ligne 7 = en-têtes tableau


# ── En-têtes tableau ──────────────────────────────────────────────────────────
def _entetes(ws, row, cols, h=18):
    ws.row_dimensions[row].height = h
    for i, col in enumerate(cols, 1):
        c = ws.cell(row=row, column=i, value=col)
        c.font      = _font(sz=10, bold=True)
        c.fill      = _fill(C_GRIS_ENT)
        c.alignment = _align("center", "center", wrap=True)
        c.border    = Border(
            top   =_S(C_BORDURE, "thin"),
            left  =_S(C_BORDURE, "thin"),
            right =_S(C_BORDURE, "thin"),
            bottom=_S(C_BORD_EP, "medium"),
        )


# ── Ligne de données ──────────────────────────────────────────────────────────
def _ligne(ws, row, data, idx=0, h=15, surbrillance=None, bold_col1=False):
    ws.row_dimensions[row].height = h
    bg = surbrillance if surbrillance else (C_GRIS_LIG if idx % 2 == 0 else C_BLANC)
    nb = len(data)
    for i, val in enumerate(data, 1):
        c = ws.cell(row=row, column=i, value=val)
        c.font      = _font(sz=10, bold=(bold_col1 and i == 1))
        c.fill      = _fill(bg)
        c.alignment = _align("left", "center", wrap=(i == nb))
        c.border    = Border(
            bottom=_S(C_BORDURE, "thin"),
            left  =_S(C_BORDURE, "thin"),
            right =_S(C_BORDURE, "thin"),
        )


# ── Ligne de titre de section ─────────────────────────────────────────────────
def _section(ws, row, texte, nb_cols, h=13):
    ws.row_dimensions[row].height = h
    for col in range(1, nb_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(C_BLANC)
        c.border = Border(bottom=_S(C_ROUGE, "thin"))
    c = ws.cell(row=row, column=1, value=texte)
    c.font      = _font(sz=10, bold=True, color=C_ROUGE)
    c.alignment = _align("left", "center")


# ── Ligne de total ────────────────────────────────────────────────────────────
def _total(ws, row, data, nb_cols, h=17):
    ws.row_dimensions[row].height = h
    for col in range(1, nb_cols + 1):
        c = ws.cell(row=row, column=col)
        c.fill   = _fill(C_GRIS_ENT)
        c.border = Border(
            top   =_S(C_BORD_EP, "medium"),
            bottom=_S(C_BORDURE, "thin"),
            left  =_S(C_BORDURE, "thin"),
            right =_S(C_BORDURE, "thin"),
        )
    for i, val in enumerate(data, 1):
        if val is None:
            continue
        c = ws.cell(row=row, column=i, value=val)
        c.font      = _font(sz=10, bold=True)
        c.alignment = _align("left", "center")


# ── Workbook helpers ──────────────────────────────────────────────────────────
def _new_wb():
    wb = Workbook()
    wb.remove(wb.active)
    return wb

def _send_wb(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"{filename}.xlsx",
    )


# ── Route principale ──────────────────────────────────────────────────────────
@app.route("/api/export/<type_export>", methods=["POST"])
def export_xlsx(type_export):
    if not OPENPYXL_OK:
        return jsonify({"erreur": "openpyxl non disponible"}), 500
    try:
        payload  = request.get_json(force=True, silent=True) or {}
        auteur   = payload.get("auteur", "—")
        bd       = payload.get("bd", {})
        data     = payload.get("data", {})
        from datetime import datetime
        date_iso = datetime.now().strftime("%Y-%m-%d")

        handlers = {
            "parc":             lambda: _export_parc(bd, auteur, date_iso),
            "entretiens":       lambda: _export_entretiens(bd, auteur, date_iso),
            "pieces":           lambda: _export_pieces(bd, auteur, date_iso),
            "pointage_jour":    lambda: _export_ptg_jour(data.get("pointages", []), auteur, date_iso),
            "pointage_semaine": lambda: _export_ptg_semaine(data.get("pointages", []), auteur, date_iso),
            "livraisons_admin": lambda: _export_liv_admin(data, auteur, date_iso),
            "livraisons_chef":  lambda: _export_liv_chef(data, auteur, date_iso),
            "total_chantier":   lambda: _export_total(data, bd, auteur, date_iso),
        }
        if type_export not in handlers:
            return jsonify({"erreur": f"Export inconnu : {type_export}"}), 400
        return handlers[type_export]()
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({"erreur": str(e)}), 500


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 1 — PARC MATÉRIEL
# A4 PAYSAGE — 10 colonnes — 107 units
# ══════════════════════════════════════════════════════════════════════════════
def _export_parc(bd, auteur, date_iso):
    machines   = bd.get("machines", [])
    categories = bd.get("categories", [])

    def ncat(cid):
        return next((c.get("nom", "—") for c in categories if c.get("id") == cid), "—")

    def statut(m):
        from datetime import date
        seuil = m.get("seuil", 250) or 250
        ecart = (m.get("heures", 0) or 0) - (m.get("hEntretien", 0) or 0)
        pct   = ecart / seuil
        vgp   = m.get("vgp", "")
        if vgp:
            try:
                days = (date.fromisoformat(vgp) - date.today()).days
                if days < 0:  return "da"
                if days < 30: return "wa"
            except:
                pass
        if pct >= 1:   return "da"
        if pct >= 0.8: return "wa"
        return "ok"

    ST_LBL  = {"da": "En retard",  "wa": "À prévoir", "ok": "À jour"}
    ST_FILL = {"da": C_ROUGE_L,    "wa": C_ORANGE_L,  "ok": None}

    #            Machine  Modèle  An  Site  H.tot H.entr Écart Seuil Statut VGP
    COLS = ["Machine", "Modèle", "An.", "Site", "H. tot.", "H. entr.", "Écart", "Seuil", "Statut", "VGP"]
    LARG = [     20,      14,     5,    12,      9,        10,         8,      8,     10,      11]  # = 107

    wb = _new_wb()
    ws = wb.create_sheet("Parc matériel")
    _set_cols(ws, LARG)
    _setup_print(ws, "landscape")

    r = _cartouche(ws, "Parc Matériel", auteur, len(COLS))
    _entetes(ws, r, COLS)
    r += 1

    cat_order = [c.get("nom", "—") for c in categories]
    cat_dict  = {}
    for m in machines:
        cn = ncat(m.get("catId", ""))
        cat_dict.setdefault(cn, []).append(m)

    idx = 0
    nb = {"da": 0, "wa": 0, "ok": 0}
    for cn in cat_order:
        ms = cat_dict.get(cn, [])
        if not ms:
            continue
        _section(ws, r, f"{cn}  ({len(ms)} engin{'s' if len(ms) > 1 else ''})", len(COLS))
        r += 1
        for m in ms:
            st    = statut(m)
            nb[st] += 1
            ecart = (m.get("heures", 0) or 0) - (m.get("hEntretien", 0) or 0)
            _ligne(ws, r, [
                m.get("nom", "—"),
                m.get("modele", "—") or "—",
                m.get("annee", "—") or "—",
                m.get("site", "—") or "—",
                f"{m.get('heures', 0)} h",
                f"{m.get('hEntretien', 0)} h",
                f"{ecart} h",
                f"{m.get('seuil', 250)} h",
                ST_LBL[st],
                _date_fr(m.get("vgp", "")) or "—",
            ], idx, surbrillance=ST_FILL[st])
            idx += 1
            r += 1

    r += 1
    _total(ws, r, [
        f"Total : {len(machines)} engin(s)  —  {nb['da']} en retard  /  {nb['wa']} à prévoir  /  {nb['ok']} à jour",
        None, None, None, None, None, None, None, None, None,
    ], len(COLS))
    return _send_wb(wb, f"FGS-Parc-{date_iso}")


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 2 — ENTRETIENS
# A4 PAYSAGE — 9 colonnes — 120 units
# ══════════════════════════════════════════════════════════════════════════════
def _export_entretiens(bd, auteur, date_iso):
    interventions = sorted(bd.get("interventions", []), key=lambda x: x.get("date", ""), reverse=True)
    machines   = bd.get("machines", [])
    pieces     = bd.get("pieces", [])
    categories = bd.get("categories", [])

    def get_m(mid): return next((m for m in machines if m.get("id") == mid), {})
    def get_p(pid): return next((p for p in pieces  if p.get("id") == pid), {})
    def ncat(cid):  return next((c.get("nom", "—") for c in categories if c.get("id") == cid), "—")

    TY_LBL  = {"entretien": "Préventif", "reparation": "Réparation",
               "VGP": "VGP", "remplacement": "Remplacement"}
    TY_FILL = {"entretien": None, "reparation": C_ROUGE_L,
               "VGP": C_BLEU_L, "remplacement": C_ORANGE_L}

    #           Date  Machine  Catégo  Site  Type   H   Pièces remplacées  Autres  Notes
    COLS = ["Date", "Machine", "Catégorie", "Site", "Type", "H.", "Pièces remplacées", "Pièces libres", "Notes"]
    LARG = [  10,     16,        11,        10,     12,     7,         20,                  16,             18]  # = 120

    wb = _new_wb()
    ws = wb.create_sheet("Entretiens")
    _set_cols(ws, LARG)
    _setup_print(ws, "landscape")

    r = _cartouche(ws, "Historique des Entretiens", auteur, len(COLS))
    _entetes(ws, r, COLS)
    r += 1

    for idx, iv in enumerate(interventions):
        m    = get_m(iv.get("machineId", ""))
        ty   = iv.get("type", "entretien")
        pcs  = [get_p(pid).get("nom", "?") for pid in (iv.get("piecesChangees") or []) if get_p(pid)]
        autr = [f"{ap.get('nom','?')} ({ap.get('ref','—')})" for ap in (iv.get("autresPieces") or [])]
        _ligne(ws, r, [
            _date_fr(iv.get("date", "")),
            m.get("nom", "—"),
            ncat(m.get("catId", "")),
            m.get("site", "—") or "—",
            TY_LBL.get(ty, ty),
            f"{iv.get('heures', '—')} h" if iv.get("heures") else "—",
            ", ".join(pcs)  if pcs  else "—",
            ", ".join(autr) if autr else "—",
            iv.get("notes", "—") or "—",
        ], idx, surbrillance=TY_FILL.get(ty))
        r += 1

    r += 1
    _total(ws, r, [f"Total : {len(interventions)} intervention(s)",
                   None, None, None, None, None, None, None, None], len(COLS))
    return _send_wb(wb, f"FGS-Entretiens-{date_iso}")


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 3 — PIÈCES & STOCKS
# A4 PAYSAGE — 9 colonnes — 107 units
# ══════════════════════════════════════════════════════════════════════════════
def _export_pieces(bd, auteur, date_iso):
    pieces   = bd.get("pieces", [])
    machines = bd.get("machines", [])

    def get_compat(p):
        ids = p.get("machinesCompatibles") or []
        if not ids:
            ids = [m.get("id") for m in machines
                   if any(pa.get("pieceId") == p.get("id") for pa in (m.get("piecesAssociees") or []))]
        return [m.get("nom", "?") for m in machines if m.get("id") in ids]

    #           Réf   Désignation  Durée  Unité  Stock  Besoin  Manque  Statut  Machines
    COLS = ["Réf.", "Désignation", "Durée", "Unité", "Stock", "Besoin", "Manque", "Statut", "Machines compatibles"]
    LARG = [  10,       20,           8,      7,       7,       7,        8,        10,             30]  # = 107

    wb = _new_wb()
    ws = wb.create_sheet("Pièces & Stocks")
    _set_cols(ws, LARG)
    _setup_print(ws, "landscape")

    r = _cartouche(ws, "Catalogue Pièces & Stocks", auteur, len(COLS))
    _entetes(ws, r, COLS)
    r += 1

    a_cmd = sorted(
        [p for p in pieces if max(len(get_compat(p)) - (p.get("stock", 0) or 0), 0) > 0],
        key=lambda p: -max(len(get_compat(p)) - (p.get("stock", 0) or 0), 0),
    )
    ok = [p for p in pieces if max(len(get_compat(p)) - (p.get("stock", 0) or 0), 0) == 0]

    if a_cmd:
        _section(ws, r, f"À commander  ({len(a_cmd)})", len(COLS))
        r += 1
        for idx, p in enumerate(a_cmd):
            compat = get_compat(p)
            stock  = p.get("stock", 0) or 0
            manque = max(len(compat) - stock, 0)
            _ligne(ws, r, [
                p.get("ref", "—"), p.get("nom", "—"),
                p.get("dureeVal", "—"), p.get("dureeUnite", "h"),
                stock, len(compat), manque, "À commander",
                ", ".join(compat) if compat else "—",
            ], idx, surbrillance=C_ROUGE_L)
            r += 1

    if ok:
        _section(ws, r, f"Stock suffisant  ({len(ok)})", len(COLS))
        r += 1
        for idx, p in enumerate(ok):
            compat = get_compat(p)
            stock  = p.get("stock", 0) or 0
            statut = "OK" if len(compat) > 0 else "Non rattachée"
            _ligne(ws, r, [
                p.get("ref", "—"), p.get("nom", "—"),
                p.get("dureeVal", "—"), p.get("dureeUnite", "h"),
                stock, len(compat) or "—", "—", statut,
                ", ".join(compat) if compat else "—",
            ], idx)
            r += 1

    r += 1
    _total(ws, r, [
        None,
        f"Total : {len(pieces)} réf.  —  {sum(p.get('stock', 0) or 0 for p in pieces)} unités en stock",
        None, None, None, None, None, None, None,
    ], len(COLS))
    return _send_wb(wb, f"FGS-Pieces-{date_iso}")


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 4 — POINTAGE JOURNALIER
# A4 PORTRAIT — 7 colonnes — 87 units
# ══════════════════════════════════════════════════════════════════════════════
def _export_ptg_jour(pointages, auteur, date_iso):
    from datetime import date
    sorted_ptg = sorted(pointages, key=lambda x: x.get("date_jour", ""), reverse=True)
    JOURS = ["Lundi", "Mardi", "Mercredi", "Jeudi", "Vendredi", "Samedi", "Dimanche"]

    #           Date   Chantier  Employé  Heures  GD     Panier  Notes
    COLS = ["Date", "Chantier", "Employé", "Heures", "GD", "Panier", "Notes"]
    LARG = [  13,      16,         15,       8,       6,     7,        22]  # = 87

    wb = _new_wb()
    ws = wb.create_sheet("Pointage journalier")
    _set_cols(ws, LARG)
    _setup_print(ws, "portrait")

    total_h = sum(float(l.get("heures", 0) or 0) for p in sorted_ptg for l in (p.get("lignes") or []))
    r = _cartouche(ws, "Pointages — Détail journalier", auteur, len(COLS))
    _entetes(ws, r, COLS)
    r += 1

    for ptg in sorted_ptg:
        dj = ptg.get("date_jour", "")
        try:
            d = date.fromisoformat(dj)
            date_lbl = f"{JOURS[d.weekday()]} {d.strftime('%d/%m/%Y')}"
        except:
            date_lbl = dj
        chantier   = ptg.get("chantier", "—")
        saisie_par = ptg.get("auteur", "—")
        lignes     = ptg.get("lignes") or []
        notes      = ptg.get("notes", "") or "—"

        _section(ws, r, f"{date_lbl}  —  {chantier}  (saisi par {saisie_par})", len(COLS))
        r += 1

        if not lignes:
            _ligne(ws, r, [date_lbl, chantier, "(aucun employé)", "—", "", "", notes], 0)
            r += 1
        else:
            for j, l in enumerate(lignes):
                surb = C_ORANGE_L if l.get("gd") else (C_VERT_L if l.get("panier") else None)
                _ligne(ws, r, [
                    date_lbl   if j == 0 else "",
                    chantier   if j == 0 else "",
                    l.get("nom", "—"),
                    l.get("heures", "—") or "—",
                    "✓" if l.get("gd")     else "",
                    "✓" if l.get("panier") else "",
                    notes      if j == 0 else "",
                ], j, surbrillance=surb)
                r += 1

        # Sous-total journée (compact, italique gris)
        th    = sum(float(l.get("heures", 0) or 0) for l in lignes)
        nb_gd = sum(1 for l in lignes if l.get("gd"))
        nb_p  = sum(1 for l in lignes if l.get("panier"))
        ws.row_dimensions[r].height = 11
        for col in range(1, len(COLS) + 1):
            ws.cell(row=r, column=col).fill = _fill(C_GRIS_LIG)
        c = ws.cell(row=r, column=1,
                    value=f"   {len(lignes)} présence(s)  ·  {th:.1f} h  ·  {nb_gd} GD  ·  {nb_p} panier(s)")
        c.font = _font(sz=8, italic=True, color="999999")
        c.alignment = _align("left", "center")
        r += 1

    r += 1
    nb_gd_tot  = sum(1 for p in sorted_ptg for l in (p.get("lignes") or []) if l.get("gd"))
    nb_pan_tot = sum(1 for p in sorted_ptg for l in (p.get("lignes") or []) if l.get("panier"))
    _total(ws, r, [
        f"Total : {len(sorted_ptg)} journée(s)  —  {total_h:.1f} h  —  {nb_gd_tot} GD  —  {nb_pan_tot} paniers",
        None, None, f"{total_h:.1f} h", str(nb_gd_tot), str(nb_pan_tot), None,
    ], len(COLS))
    return _send_wb(wb, f"FGS-Pointage-Jour-{date_iso}")


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 5 — POINTAGE HEBDOMADAIRE  (une feuille par semaine)
# A4 PAYSAGE — 12 colonnes — 104 units
# ══════════════════════════════════════════════════════════════════════════════
def _export_ptg_semaine(pointages, auteur, date_iso):
    from datetime import date, timedelta

    sem_map = {}
    for ptg in pointages:
        dj = ptg.get("date_jour", "")
        try:
            d   = date.fromisoformat(dj)
            lun = d - timedelta(days=d.weekday())
            wk  = lun.isoformat()
        except:
            wk = dj[:7]
        sem_map.setdefault(wk, []).append(ptg)

    JOURS = ["Lun", "Mar", "Mer", "Jeu", "Ven", "Sam", "Dim"]
    #          Employé  Chantier  L  M  M  J  V  S  D  Th  GD  Pan
    COLS = ["Employé", "Chantier"] + JOURS + ["Total h", "GD", "Pan."]
    LARG = [    17,        15,     7, 7, 7, 7, 7, 7, 7,   9,   6,   8]  # = 104

    wb = _new_wb()
    for wk in sorted(sem_map.keys(), reverse=True):
        ptgs = sem_map[wk]
        try:
            lun = date.fromisoformat(wk)
            dim = lun + timedelta(days=6)
            lbl = f"Semaine du {lun.strftime('%d/%m')} au {dim.strftime('%d/%m/%Y')}"
            nom_feuille = f"Sem {lun.strftime('%d.%m.%y')}"[:31]
        except:
            lbl = wk
            nom_feuille = wk[:31]

        ws = wb.create_sheet(nom_feuille)
        _set_cols(ws, LARG)
        _setup_print(ws, "landscape")

        r = _cartouche(ws, f"Pointages hebdomadaires — {lbl}", auteur, len(COLS))
        _entetes(ws, r, COLS)
        r += 1

        emp_map = {}
        for ptg in ptgs:
            dj = ptg.get("date_jour", "")
            try:
                dow = date.fromisoformat(dj).weekday()
            except:
                dow = 0
            for l in (ptg.get("lignes") or []):
                k = f"{l.get('nom','?')}||{ptg.get('chantier','?')}"
                if k not in emp_map:
                    emp_map[k] = {"nom": l.get("nom", "?"), "chantier": ptg.get("chantier", "?"),
                                  "h": [0.0] * 7, "gd": 0, "pan": 0}
                emp_map[k]["h"][dow] += float(l.get("heures", 0) or 0)
                if l.get("gd"):     emp_map[k]["gd"]  += 1
                if l.get("panier"): emp_map[k]["pan"] += 1

        idx = 0
        for ch in sorted(set(v["chantier"] for v in emp_map.values())):
            emps = [v for v in emp_map.values() if v["chantier"] == ch]
            _section(ws, r, ch, len(COLS))
            r += 1
            for e in sorted(emps, key=lambda x: x["nom"]):
                tot  = sum(e["h"])
                surb = C_ORANGE_L if (e["gd"] or e["pan"]) else None
                _ligne(ws, r, (
                    [e["nom"], e["chantier"]] +
                    [round(h, 1) if h > 0 else "" for h in e["h"]] +
                    [round(tot, 1), e["gd"] or "", e["pan"] or ""]
                ), idx, surbrillance=surb)
                idx += 1
                r += 1

        h_jours  = [sum(v["h"][i] for v in emp_map.values()) for i in range(7)]
        tot_sem  = sum(h_jours)
        tot_gd   = sum(v["gd"]  for v in emp_map.values())
        tot_pan  = sum(v["pan"] for v in emp_map.values())
        r += 1
        _total(ws, r, (
            [f"Total semaine : {tot_sem:.1f} h", ""] +
            [round(h, 1) if h > 0 else "" for h in h_jours] +
            [round(tot_sem, 1), tot_gd or "", tot_pan or ""]
        ), len(COLS))

    return _send_wb(wb, f"FGS-Pointage-Semaine-{date_iso}")


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 6 — LIVRAISONS ADMIN  (2 feuilles)
# Récap : A4 PORTRAIT — 6 colonnes — 83 units
# Détail : A4 PAYSAGE  — 9 colonnes — 113 units
# ══════════════════════════════════════════════════════════════════════════════
def _export_liv_admin(data, auteur, date_iso):
    livraisons = data.get("livraisons", [])
    chantier   = data.get("chantier", "Tous chantiers") or "Tous chantiers"
    filtre_sem = data.get("semaine", "")
    if not livraisons:
        return jsonify({"erreur": "Aucune livraison à exporter"}), 400

    wb    = _new_wb()
    sous_t = chantier + (f" — Semaine {filtre_sem}" if filtre_sem else "")

    # ── Feuille 1 : Récapitulatif (portrait) ──────────────────────────────────
    #              Matériau  Quantité  Unité  Nb liv.  Prix unit.  Coût HT
    COLS_R = ["Matériau", "Quantité", "Unité", "Nb liv.", "Prix unit.", "Coût total HT"]
    LARG_R = [     26,        13,       8,       9,          13,            14]  # = 83
    ws1 = wb.create_sheet("Récapitulatif")
    _set_cols(ws1, LARG_R)
    _setup_print(ws1, "portrait")
    r = _cartouche(ws1, f"Livraisons — {sous_t}", auteur, len(COLS_R))
    _entetes(ws1, r, COLS_R)
    r += 1

    agg = {}
    for l in livraisons:
        el = l.get("element", "?")
        if el not in agg:
            agg[el] = {"element": el, "unite": l.get("unite", ""), "qte": 0.0, "nb": 0, "prix": None}
        agg[el]["qte"] += float(l.get("quantite", 0) or 0)
        agg[el]["nb"]  += 1
        if l.get("prix_unitaire") and not agg[el]["prix"]:
            agg[el]["prix"] = float(l["prix_unitaire"])

    total_cout = 0.0
    for idx, v in enumerate(sorted(agg.values(), key=lambda x: x["qte"], reverse=True)):
        prix = v.get("prix")
        cout = round(v["qte"] * prix, 2) if prix else None
        if cout: total_cout += cout
        _ligne(ws1, r, [
            v["element"], round(v["qte"], 3), v["unite"], v["nb"],
            f"{prix:.2f} €" if prix else "—",
            f"{cout:.2f} €" if cout else "—",
        ], idx)
        r += 1

    r += 1
    _total(ws1, r, [
        f"Total : {len(livraisons)} livraison(s)", None, None, len(livraisons), None,
        f"{total_cout:.2f} €" if total_cout else "—",
    ], len(COLS_R))

    # ── Feuille 2 : Détail (paysage) ─────────────────────────────────────────
    #               Date  Chantier  Saisi par  Matériau  Qté  Unité  Prix  Coût  Notes
    COLS_D = ["Date", "Chantier", "Saisi par", "Matériau", "Qté", "Unité", "Prix", "Coût HT", "Notes"]
    LARG_D = [  10,     16,          13,           18,      9,      7,      11,      11,         18]  # = 113
    ws2 = wb.create_sheet("Détail livraisons")
    _set_cols(ws2, LARG_D)
    _setup_print(ws2, "landscape")
    r = _cartouche(ws2, f"Livraisons — Détail  ({sous_t})", auteur, len(COLS_D))
    _entetes(ws2, r, COLS_D)
    r += 1

    for idx, l in enumerate(sorted(livraisons, key=lambda x: x.get("date_liv", ""), reverse=True)):
        prix = l.get("prix_unitaire")
        prix = float(prix) if prix else None
        cout = round(float(l.get("quantite", 0) or 0) * prix, 2) if prix else None
        _ligne(ws2, r, [
            _date_fr(l.get("date_liv", "")), l.get("chantier", "—"), l.get("auteur", "—"),
            l.get("element", "—"), round(float(l.get("quantite", 0) or 0), 3),
            l.get("unite", "—"),
            f"{prix:.2f} €" if prix else "—",
            f"{cout:.2f} €" if cout else "—",
            l.get("notes", "—") or "—",
        ], idx)
        r += 1

    r += 1
    _total(ws2, r, [
        f"Total : {len(livraisons)} livraison(s)",
        None, None, None, None, None, None,
        f"{total_cout:.2f} €" if total_cout else "—", None,
    ], len(COLS_D))

    nom = chantier.replace(" ", "-").replace("/", "-")[:20]
    return _send_wb(wb, f"FGS-Livraisons-{nom}-{date_iso}")


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 7 — LIVRAISONS CHEF
# A4 PORTRAIT — 6 colonnes — 84 units
# ══════════════════════════════════════════════════════════════════════════════
def _export_liv_chef(data, auteur, date_iso):
    livraisons = data.get("livraisons", [])
    if not livraisons:
        return jsonify({"erreur": "Aucune livraison à exporter"}), 400

    #           Date  Chantier  Matériau  Quantité  Unité  Notes
    COLS = ["Date", "Chantier", "Matériau", "Quantité", "Unité", "Notes"]
    LARG = [  12,      20,         22,         10,        9,       11]  # = 84

    wb = _new_wb()
    ws = wb.create_sheet("Mes livraisons")
    _set_cols(ws, LARG)
    _setup_print(ws, "portrait")

    r = _cartouche(ws, "Mes Livraisons Chantier", auteur, len(COLS))
    _entetes(ws, r, COLS)
    r += 1

    for idx, l in enumerate(sorted(livraisons, key=lambda x: x.get("date_liv", ""), reverse=True)):
        _ligne(ws, r, [
            _date_fr(l.get("date_liv", "")), l.get("chantier", "—"),
            l.get("element", "—"), round(float(l.get("quantite", 0) or 0), 3),
            l.get("unite", "—"), l.get("notes", "—") or "—",
        ], idx)
        r += 1

    r += 1
    _total(ws, r, [f"Total : {len(livraisons)} livraison(s)", None, None, None, None, None], len(COLS))
    return _send_wb(wb, f"FGS-MesLivraisons-{date_iso}")


# ══════════════════════════════════════════════════════════════════════════════
# EXPORT 8 — TOTAL CHANTIER  (3 feuilles)
# Synthèse   : A4 PORTRAIT  — 4 col — 83 units
# Matériaux  : A4 PORTRAIT  — 5 col — 76 units
# Pointages  : A4 PAYSAGE   — 7 col — 92 units
# ══════════════════════════════════════════════════════════════════════════════
def _export_total(data, bd, auteur, date_iso):
    chantier   = data.get("chantier", "—")
    pointages  = data.get("pointages", [])
    livraisons = data.get("livraisons", [])
    prix_ref   = data.get("prix_ref", {})

    total_h = sum(float(l.get("heures", 0) or 0) for p in pointages for l in (p.get("lignes") or []))
    nb_pres = sum(len(p.get("lignes") or []) for p in pointages)
    nb_gd   = sum(1 for p in pointages for l in (p.get("lignes") or []) if l.get("gd"))
    nb_pan  = sum(1 for p in pointages for l in (p.get("lignes") or []) if l.get("panier"))
    taux_h  = prix_ref.get("__heures__")
    cout_mo = round(total_h * float(taux_h), 2) if taux_h else None

    agg_mat = {}
    for l in livraisons:
        el = l.get("element", "?")
        if el not in agg_mat:
            agg_mat[el] = {"qte": 0.0, "unite": l.get("unite", ""), "prix": None}
        agg_mat[el]["qte"] += float(l.get("quantite", 0) or 0)
        if l.get("prix_unitaire") and not agg_mat[el]["prix"]:
            agg_mat[el]["prix"] = float(l["prix_unitaire"])

    total_mat  = 0.0
    mat_lignes = []
    for el, v in agg_mat.items():
        prix = float(prix_ref[el]) if el in prix_ref else v.get("prix")
        cout = round(v["qte"] * prix, 2) if prix else None
        if cout: total_mat += cout
        mat_lignes.append((el, round(v["qte"], 3), v["unite"], prix, cout))

    total_gen = (cout_mo or 0) + total_mat
    wb = _new_wb()

    # ── Feuille 1 : Synthèse (portrait) ──────────────────────────────────────
    #              Poste  Détail  Taux  Montant HT
    COLS_S = ["Poste", "Détail", "Taux", "Montant HT"]
    LARG_S = [  16,      42,      12,       13]  # = 83
    ws1 = wb.create_sheet("Synthèse")
    _set_cols(ws1, LARG_S)
    _setup_print(ws1, "portrait")
    r = _cartouche(ws1, f"Synthèse chantier — {chantier}", auteur, len(COLS_S))
    _entetes(ws1, r, COLS_S)
    r += 1
    _ligne(ws1, r, [
        "Main d'œuvre",
        f"{total_h:.1f} h · {nb_pres} présence(s) · {nb_gd} GD · {nb_pan} panier(s)",
        f"{taux_h} €/h" if taux_h else "—",
        f"{cout_mo:.2f} €" if cout_mo else "—",
    ], 0, bold_col1=True)
    r += 1
    _ligne(ws1, r, [
        "Matériaux",
        ", ".join(agg_mat.keys()) or "—",
        "→ Feuille Matériaux",
        f"{total_mat:.2f} €" if total_mat else "—",
    ], 1, bold_col1=True)
    r += 2
    _total(ws1, r, ["TOTAL GÉNÉRAL", None, None, f"{total_gen:.2f} €" if total_gen else "—"], len(COLS_S))

    # ── Feuille 2 : Matériaux (portrait) ─────────────────────────────────────
    #               Matériau  Qté   Unité  Prix unit. HT  Coût HT
    COLS_M = ["Matériau", "Quantité", "Unité", "Prix unit. HT", "Coût total HT"]
    LARG_M = [    26,        12,        8,           16,               14]  # = 76
    ws2 = wb.create_sheet("Matériaux")
    _set_cols(ws2, LARG_M)
    _setup_print(ws2, "portrait")
    r = _cartouche(ws2, f"Matériaux — {chantier}", auteur, len(COLS_M))
    _entetes(ws2, r, COLS_M)
    r += 1
    for idx, (el, qte, unite, prix, cout) in enumerate(sorted(mat_lignes, key=lambda x: -(x[4] or 0))):
        _ligne(ws2, r, [
            el, qte, unite,
            f"{prix:.2f} €/{unite}" if prix else "—",
            f"{cout:.2f} €" if cout else "—",
        ], idx)
        r += 1
    r += 1
    _total(ws2, r, [
        f"Total : {len(mat_lignes)} matériau(x)", None, None, None,
        f"{total_mat:.2f} €" if total_mat else "—",
    ], len(COLS_M))

    # ── Feuille 3 : Pointages (paysage) ──────────────────────────────────────
    #               Date  Saisi par  Employés présents  Présences  Heures  GD  Paniers
    COLS_P = ["Date", "Saisi par", "Employés présents", "Présences", "Heures", "GD", "Paniers"]
    LARG_P = [  10,      14,               34,              10,         9,      7,       8]  # = 92
    ws3 = wb.create_sheet("Pointages")
    _set_cols(ws3, LARG_P)
    _setup_print(ws3, "landscape")
    r = _cartouche(ws3, f"Pointages — {chantier}", auteur, len(COLS_P))
    _entetes(ws3, r, COLS_P)
    r += 1
    for idx, ptg in enumerate(sorted(pointages, key=lambda x: x.get("date_jour", ""), reverse=True)):
        lignes = ptg.get("lignes") or []
        th  = sum(float(l.get("heures", 0) or 0) for l in lignes)
        ngd = sum(1 for l in lignes if l.get("gd"))
        npa = sum(1 for l in lignes if l.get("panier"))
        _ligne(ws3, r, [
            _date_fr(ptg.get("date_jour", "")),
            ptg.get("auteur", "—"),
            ", ".join(l.get("nom", "?") for l in lignes) or "—",
            len(lignes), round(th, 1),
            ngd or "—", npa or "—",
        ], idx)
        r += 1
    r += 1
    _total(ws3, r, [
        f"Total : {len(pointages)} journée(s)", None, None,
        nb_pres, round(total_h, 1), nb_gd or "—", nb_pan or "—",
    ], len(COLS_P))

    nom = chantier.replace(" ", "-").replace("/", "-")[:20]
    return _send_wb(wb, f"FGS-Total-{nom}-{date_iso}")
